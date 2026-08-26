"""Экспорт свода по пробежке в .xlsx — два листа: «Бегуны» и «Волонтёры».

Первый экспорт таблиц на сайте. Файл собирается в память из готового payload
build_event_svod (без повторных запросов к БД): скачивание идёт сразу после
просмотра той же таблицы на странице.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def svod_export_filename(slug: str, event_date: str) -> str:
    return f"svod_{slug}_{event_date}.xlsx"


def _yes(value: object) -> str:
    return "да" if value else ""


def _fill_sheet(
    sheet: Worksheet,
    headers: list[tuple[str, int]],
    rows: list[list[Any]],
) -> None:
    header_font = Font(bold=True)
    for column, (title, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column, value=value)
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        )


def build_svod_workbook(payload: dict[str, Any]) -> bytes:
    event = payload["event"]
    workbook = Workbook()

    runners_sheet = workbook.active
    runners_sheet.title = "Бегуны"
    runner_headers = [
        ("Позиция", 9),
        ("Имя", 28),
        ("Время", 9),
        ("Новичок системы", 12),
        ("Впервые на локации", 12),
        ("Личный рекорд", 10),
        ("Рекорд на локации", 10),
        ("Вернулся после паузы >года", 12),
        ("Пробежек здесь", 10),
        ("Пробежек в системе", 10),
        ("Юбилей на локации", 10),
        ("Следующая — юбилейная здесь", 12),
        ("Юбилей в системе", 10),
        ("Следующая — юбилейная в системе", 12),
    ]
    runner_rows = [
        [
            runner.get("position"),
            runner.get("name") or "—",
            runner.get("finish_time_display"),
            _yes(runner.get("first_in_system")),
            _yes(runner.get("first_at_location") and not runner.get("first_in_system")),
            _yes(runner.get("is_pb")),
            _yes(runner.get("is_location_pb")),
            _yes(runner.get("comeback")),
            runner.get("location_runs_count"),
            runner.get("platform_runs_count"),
            runner.get("location_milestone") or "",
            runner.get("location_next_milestone") or "",
            runner.get("platform_milestone") or "",
            runner.get("platform_next_milestone") or "",
        ]
        for runner in payload["runners"]
    ]
    _fill_sheet(runners_sheet, runner_headers, runner_rows)

    volunteers_sheet = workbook.create_sheet("Волонтёры")
    volunteer_headers = [
        ("Имя", 28),
        ("Роли (волонтёрств в роли)", 34),
        ("Новые роли", 24),
        ("Впервые волонтёрит", 12),
        ("Впервые на локации", 12),
        ("Волонтёрств здесь", 10),
        ("Волонтёрств в системе", 10),
        ("Юбилей на локации", 10),
        ("Следующее — юбилейное здесь", 12),
        ("Юбилей в системе", 10),
        ("Юбилей в роли", 18),
    ]
    volunteer_rows = []
    for volunteer in payload["volunteers"]:
        roles = volunteer.get("roles") or []
        role_jubilees = [
            f"{role['label']}: {role['milestone']}" for role in roles if role.get("milestone")
        ]
        volunteer_rows.append(
            [
                volunteer.get("name") or "—",
                ", ".join(f"{role['label']} ({role['count']})" for role in roles),
                ", ".join(volunteer.get("new_roles") or []),
                _yes(volunteer.get("first_volunteering")),
                _yes(
                    volunteer.get("first_at_location")
                    and not volunteer.get("first_volunteering")
                ),
                volunteer.get("location_vol_count"),
                volunteer.get("platform_vol_count"),
                volunteer.get("location_milestone") or "",
                volunteer.get("location_next_milestone") or "",
                volunteer.get("platform_milestone") or "",
                ", ".join(role_jubilees),
            ]
        )
    _fill_sheet(volunteers_sheet, volunteer_headers, volunteer_rows)

    info_sheet = workbook.create_sheet("О файле")
    info_rows = [
        ("Локация", event.get("location_name")),
        ("Дата события", str(event.get("event_date"))),
        ("Номер события", event.get("event_number")),
        ("Система", event.get("platform_name")),
        ("Финишёров", event.get("finishers_count")),
        ("Волонтёров", event.get("volunteers_count")),
        ("Источник", "Кабинет организатора run5k.run"),
    ]
    for row_index, (label, value) in enumerate(info_rows, start=1):
        info_sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        info_sheet.cell(row=row_index, column=2, value=value)
    info_sheet.column_dimensions["A"].width = 16
    info_sheet.column_dimensions["B"].width = 40

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

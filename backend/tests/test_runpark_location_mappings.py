from __future__ import annotations

from pathlib import Path

import pytest

from app.runpark.mappings import (
    RunparkMappingDecision,
    classify_decision,
    parse_bool_flag,
    parse_duplicate_match,
    parse_runpark_slug,
    parse_xlsx_row,
    runpark_protocol_base,
    runpark_protocol_url,
)


def test_parse_bool_flag() -> None:
    assert parse_bool_flag(1) is True
    assert parse_bool_flag(0) is False
    assert parse_bool_flag("да") is True


def test_parse_runpark_slug() -> None:
    assert parse_runpark_slug("https://runpark.ru/Places/beguglich") == "beguglich"


def test_runpark_protocol_base_prefers_public_url() -> None:
    # public_url is authoritative — the slug there can diverge from the location key.
    assert (
        runpark_protocol_base("https://runpark.ru/Places/magnitprityagenie/", None)
        == "https://runpark.ru/Places/magnitprityagenie"
    )


def test_runpark_protocol_base_falls_back_to_slug() -> None:
    assert runpark_protocol_base(None, "buenosaires") == "https://runpark.ru/Places/buenosaires"
    assert runpark_protocol_base("", "") is None
    assert runpark_protocol_base(None, None) is None


def test_runpark_protocol_url() -> None:
    assert (
        runpark_protocol_url("https://runpark.ru/Places/buenosaires", None, 52)
        == "https://runpark.ru/Places/buenosaires/event/52"
    )
    assert runpark_protocol_url("https://runpark.ru/Places/buenosaires", None, None) is None
    assert runpark_protocol_url(None, None, 52) is None


def test_parse_duplicate_match_five_verst() -> None:
    platform, slug = parse_duplicate_match("five_verst: sosnovka / Сосновка (0.02 km)")
    assert platform == "five_verst"
    assert slug == "sosnovka"


def test_classify_decision_remove() -> None:
    decision, show_on_map, transitional, _ = classify_decision("Из runpark удаляем локацию")
    assert decision == RunparkMappingDecision.REMOVE_FROM_RUNPARK
    assert show_on_map is False
    assert transitional is False


def test_classify_decision_keep() -> None:
    decision, show_on_map, _, display_name = classify_decision('Оставляем с названием "#БегУглич"')
    assert decision == RunparkMappingDecision.KEEP_RUNPARK
    assert show_on_map is True
    assert display_name == "#БегУглич"


def test_classify_decision_match_parkrun() -> None:
    text = 'Название локации фиксируем "Лесопарк Северный" и мэтчим с локацией parkrun'
    decision, show_on_map, _, display_name = classify_decision(text)
    assert decision == RunparkMappingDecision.MATCH_PARKRUN
    assert show_on_map is True
    assert display_name == "Лесопарк Северный"


def test_classify_decision_transitional() -> None:
    text = (
        "Пока под вопросом, эта локация была в паркране (Жуковский), "
        "потом была вот здесь в runpark и перешла в 5 верст"
    )
    decision, show_on_map, transitional, _ = classify_decision(text)
    assert decision == RunparkMappingDecision.TRANSITIONAL
    assert show_on_map is False
    assert transitional is True


def test_parse_xlsx_row_from_fixture() -> None:
    row = (
        "397A2E1D-DEBA-483F-826D-469E273B56CD",
        "г. Углич, набережная Волги",
        None,
        'Оставляем с названием "#БегУглич"',
        "Углич",
        None,
        None,
        "57.5361",
        "38.3238",
        1,
        1,
        "https://runpark.ru/Places/beguglich",
        None,
        None,
    )
    entry = parse_xlsx_row(row, source_batch="test_batch")
    assert entry.runpark_location_id == "397A2E1D-DEBA-483F-826D-469E273B56CD"
    assert entry.runpark_slug == "beguglich"
    assert entry.display_name == "#БегУглич"
    assert entry.show_on_map is True
    assert entry.is_paused is True
    assert entry.decision == RunparkMappingDecision.KEEP_RUNPARK


@pytest.mark.parametrize(
    ("xlsx_name", "expected_count"),
    [
        ("vw_locations_202605271905_matched.xlsx", 51),
    ],
)
def test_load_reviewed_xlsx_row_count(xlsx_name: str, expected_count: int) -> None:
    path = Path(__file__).resolve().parents[2] / "data" / "runpark_import" / xlsx_name
    if not path.is_file():
        pytest.skip("reviewed xlsx not present")
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        pytest.skip("openpyxl not installed")

    from app.runpark.mappings import parse_xlsx_row

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    entries = [
        parse_xlsx_row(row, source_batch=path.stem)
        for row in ws.iter_rows(min_row=3, values_only=True)
        if row and row[0]
    ]
    wb.close()
    assert len(entries) == expected_count
    assert sum(1 for entry in entries if entry.show_on_map) == 11
    assert sum(1 for entry in entries if entry.is_transitional) == 2

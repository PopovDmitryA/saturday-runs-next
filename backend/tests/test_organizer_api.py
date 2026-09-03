"""Кабинет организатора: доступ, «Долгая пауза», свод и экспорт."""

from __future__ import annotations

from collections.abc import Generator
from datetime import date, timedelta
from uuid import uuid4

import fakeredis
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.config import Settings, get_settings
from app.db.session import get_db
from app.main import app
from app.models import (
    Event,
    Location,
    LocationDescription,
    LocationOrganizerAccess,
    Participant,
    Platform,
    PlatformLink,
    ProtocolRevision,
    ProtocolUploadFact,
    RunResult,
    User,
    VolunteerResult,
)
from app.services.location_page_service import LOCATIONS_INDEX_CACHE_KEY
from app.services.organizer_access_service import (
    derive_organizer_identity_keys,
    invalidate_organizer_locations_cache,
)

ADMIN_TELEGRAM_ID = 9101


@pytest.fixture
def auth_settings() -> Settings:
    return Settings(
        app_secret_key="test-secret-key",
        app_debug=True,
        telegram_bot_internal_secret="bot-secret",
        telegram_bot_username="TestBot",
        admin_telegram_id=ADMIN_TELEGRAM_ID,
        database_url=get_settings().database_url,
        redis_url="redis://localhost:6379/0",
    )


@pytest.fixture
def client(
    db_session: Session, fake_redis: fakeredis.FakeRedis, auth_settings: Settings
) -> Generator[TestClient, None, None]:
    def override_get_db() -> Generator[Session, None, None]:
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[get_settings] = lambda: auth_settings

    with TestClient(app) as test_client:
        yield test_client

    app.dependency_overrides.clear()


def _login(client: TestClient, telegram_id: int, username: str = "user") -> None:
    request_token = client.post("/api/auth/login-request").json()["request_token"]
    confirm = client.post(
        "/api/auth/bot/confirm",
        json={
            "request_token": request_token,
            "telegram_id": telegram_id,
            "telegram_username": username,
            "telegram_chat_id": telegram_id,
            "consent_accepted": True,
        },
        headers={"X-Bot-Secret": "bot-secret"},
    )
    token = confirm.json()["magic_link"].split("token=")[1]
    client.get("/api/auth/callback", params={"token": token}, follow_redirects=False)


def _platform(db_session: Session, code: str, name: str) -> Platform:
    platform = db_session.query(Platform).filter(Platform.code == code).one_or_none()
    if platform is None:
        platform = Platform(code=code, name=name)
        db_session.add(platform)
        db_session.flush()
    return platform


def _location(db_session: Session, platform: Platform, slug: str) -> Location:
    location = Location(
        platform_id=platform.id,
        external_key=slug,
        name=f"Локация {slug}",
        city="Москва",
        country="Россия",
    )
    db_session.add(location)
    db_session.flush()
    return location


def _event(
    db_session: Session,
    platform: Platform,
    location: Location,
    event_date: date,
    number: int,
    *,
    is_test: bool = False,
) -> Event:
    event = Event(
        platform_id=platform.id,
        location_id=location.id,
        external_event_key=f"org-{location.external_key}-{number}",
        event_date=event_date,
        event_number=number,
        title=f"Событие {number}",
        is_test_event=is_test,
    )
    db_session.add(event)
    db_session.flush()
    return event


def _participant(db_session: Session, platform: Platform, suffix: str, name: str) -> Participant:
    participant = Participant(
        platform_id=platform.id,
        external_user_id=f"org-{suffix}",
        display_name=name,
        profile_url=f"https://example.test/{suffix}/",
    )
    db_session.add(participant)
    db_session.flush()
    return participant


def _link(db_session: Session, user: User, platform: Platform, participant: Participant) -> None:
    db_session.add(
        PlatformLink(
            user_id=user.id,
            platform_id=platform.id,
            participant_id=participant.id,
            external_user_id=participant.external_user_id,
            external_url=participant.profile_url,
        )
    )
    db_session.flush()


def _run(
    db_session: Session,
    event: Event,
    participant: Participant,
    *,
    finish_time_sec: int | None = 1200,
    position: int | None = None,
) -> None:
    db_session.add(
        RunResult(
            event_id=event.id,
            participant_id=participant.id,
            external_result_key=f"run-{uuid4()}",
            position=position,
            finish_time_sec=finish_time_sec,
            finish_time_display="00:20:00" if finish_time_sec else None,
            status="finished",
        )
    )


def _volunteer(db_session: Session, event: Event, participant: Participant, role: str) -> None:
    db_session.add(
        VolunteerResult(
            event_id=event.id,
            participant_id=participant.id,
            external_result_key=f"vol-{uuid4()}",
            role=role,
        )
    )


def _current_user(db_session: Session, telegram_id: int) -> User:
    return db_session.query(User).filter(User.telegram_id == telegram_id).one()


# ===== Автодоступ по роли организатора =====


@pytest.mark.parametrize(
    "platform_code,platform_name,role",
    [
        ("five_verst", "5 вёрст", "Организатор"),
        ("runpark", "Runpark", "Руководитель"),
        # parkrun приклеивает к роли счётчик кредитов — сырое сравнение его не поймает.
        ("parkrun", "parkrun", "Run Director (12×)"),
        ("s95", "с95", "Директор 25"),
    ],
)
def test_organizer_role_aliases_grant_access(
    db_session: Session, platform_code: str, platform_name: str, role: str
) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, platform_code, platform_name)
    location = _location(db_session, platform, f"org-loc-{suffix}")
    event = _event(db_session, platform, location, date(2026, 3, 7), 1)
    participant = _participant(db_session, platform, suffix, "Организатор Иванов")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, participant)
    _volunteer(db_session, event, participant, role)
    db_session.commit()

    keys = derive_organizer_identity_keys(db_session, user.id)
    assert len(keys) == 1


def test_non_organizer_volunteer_has_no_access(db_session: Session) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    event = _event(db_session, platform, location, date(2026, 3, 7), 1)
    participant = _participant(db_session, platform, suffix, "Волонтёр Петров")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, participant)
    _volunteer(db_session, event, participant, "Секундомер")
    db_session.commit()

    assert derive_organizer_identity_keys(db_session, user.id) == set()


def test_test_events_do_not_grant_access(db_session: Session) -> None:
    """Организатор тестового события кабинет не открывает."""
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    event = _event(db_session, platform, location, date(2026, 3, 7), 1, is_test=True)
    participant = _participant(db_session, platform, suffix, "Организатор Теста")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, participant)
    _volunteer(db_session, event, participant, "Организатор")
    db_session.commit()

    assert derive_organizer_identity_keys(db_session, user.id) == set()


# ===== API: гейт доступа =====


def test_absence_forbidden_without_access(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    _event(db_session, platform, location, date(2026, 3, 7), 1)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)

    telegram_id = int(uuid4().int % 10_000_000_000)
    _login(client, telegram_id, "stranger")

    response = client.get(f"/api/organizer/{location.external_key}/absence")
    assert response.status_code == 403


def test_admin_passes_to_any_location(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    _event(db_session, platform, location, date(2026, 3, 7), 1)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)

    _login(client, ADMIN_TELEGRAM_ID, "admin_user")

    response = client.get(f"/api/organizer/{location.external_key}/absence")
    assert response.status_code == 200


def test_manual_grant_opens_access_and_menu_flag(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    _event(db_session, platform, location, date(2026, 3, 7), 1)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)

    telegram_id = int(uuid4().int % 10_000_000_000)
    _login(client, telegram_id, "granted")
    user = _current_user(db_session, telegram_id)

    assert client.get("/api/auth/me").json()["is_organizer"] is False
    assert client.get(f"/api/organizer/{location.external_key}/absence").status_code == 403

    from app.services.location_catalog_service import LocationCatalogIndex

    identity_key = LocationCatalogIndex(db_session).canonical_identity_key(location, "five_verst")
    db_session.add(LocationOrganizerAccess(user_id=user.id, location_key=identity_key))
    db_session.commit()
    invalidate_organizer_locations_cache(user.id)

    assert client.get(f"/api/organizer/{location.external_key}/absence").status_code == 200
    assert client.get("/api/auth/me").json()["is_organizer"] is True
    listing = client.get("/api/organizer/locations").json()
    assert listing["total"] == 1
    assert listing["items"][0]["access_source"] == "manual"


def test_admin_grant_crud_invalidates_cache(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    _event(db_session, platform, location, date(2026, 3, 7), 1)
    target = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"t{suffix}")
    db_session.add(target)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)

    _login(client, ADMIN_TELEGRAM_ID, "admin_user")

    from app.services.location_catalog_service import LocationCatalogIndex

    identity_key = LocationCatalogIndex(db_session).canonical_identity_key(location, "five_verst")

    created = client.post(
        f"/api/admin/users/{target.id}/organizer-access",
        json={"location_key": identity_key, "note": "оргкоманда"},
    )
    assert created.status_code == 201
    assert len(created.json()["manual"]) == 1

    grant_id = created.json()["manual"][0]["id"]
    deleted = client.delete(f"/api/admin/users/{target.id}/organizer-access/{grant_id}")
    assert deleted.status_code == 200
    assert deleted.json()["manual"] == []


def test_admin_grant_rejects_unknown_location_key(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    target = User(
        telegram_id=int(uuid4().int % 10_000_000_000),
        telegram_username=f"t{uuid4().int % 1_000_000}",
    )
    db_session.add(target)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)

    _login(client, ADMIN_TELEGRAM_ID, "admin_user")
    response = client.post(
        f"/api/admin/users/{target.id}/organizer-access",
        json={"location_key": "no-such-location-key"},
    )
    assert response.status_code == 404


# ===== «Долгая пауза» =====


def _absence_fixture(db_session: Session) -> tuple[Location, Platform, User]:
    """Локация с 6 событиями: «пропавший» бегал первые 3, активный — все."""
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    events = [
        _event(db_session, platform, location, date(2026, 1, 3) + timedelta(days=7 * i), i + 1)
        for i in range(6)
    ]
    lost = _participant(db_session, platform, f"{suffix}-lost", "Пропавший Иванов")
    active = _participant(db_session, platform, f"{suffix}-active", "Активный Петров")
    for event in events[:3]:
        _run(db_session, event, lost)
    for event in events:
        _run(db_session, event, active)

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор Сидоров")
    _volunteer(db_session, events[0], organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    return location, platform, user


def test_absence_thresholds(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    location, _platform_obj, user = _absence_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    # Порог в 3 пробежки: «пропавший» (3 пробежки, 3 пропущенных события) виден.
    response = client.get(
        f"/api/organizer/{location.external_key}/absence",
        params={"min_runs": 3, "min_missed": 1},
    )
    assert response.status_code == 200
    payload = response.json()
    names = {item["name"] for item in payload["items"]}
    assert "Пропавший Иванов" in names
    # Активный бегал на последнем событии — пропущенных нет, в выдачу не идёт.
    assert "Активный Петров" not in names
    lost_row = next(item for item in payload["items"] if item["name"] == "Пропавший Иванов")
    assert lost_row["runs_here"] == 3
    assert lost_row["missed_events"] == 3

    # Порог по пробежкам выше — не попадает никто.
    stricter = client.get(
        f"/api/organizer/{location.external_key}/absence",
        params={"min_runs": 5, "min_missed": 1},
    )
    assert stricter.json()["total"] == 0

    # Порог по пропускам выше числа пропущенных — тоже пусто.
    stricter_missed = client.get(
        f"/api/organizer/{location.external_key}/absence",
        params={"min_runs": 3, "min_missed": 4},
    )
    assert stricter_missed.json()["total"] == 0


def test_absence_marks_who_still_runs_elsewhere(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Ушёл на соседнюю площадку — дата и подсказка; никуда не ушёл — пусто."""
    location, platform, user = _absence_fixture(db_session)
    suffix = str(uuid4().int % 1_000_000)
    other = _location(db_session, platform, f"org-other-{suffix}")
    later = _event(db_session, platform, other, date(2026, 6, 6), 1)

    lost = (
        db_session.query(Participant)
        .filter(Participant.display_name == "Пропавший Иванов")
        .order_by(Participant.created_at.desc())
        .first()
    )
    _run(db_session, later, lost)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/absence",
        params={"min_runs": 3, "min_missed": 1},
    )
    assert response.status_code == 200
    row = next(item for item in response.json()["items"] if item["name"] == "Пропавший Иванов")
    assert row["elsewhere_date_display"] == "06.06.2026"
    assert row["elsewhere_hint"] == f"{other.name} — пробежка"


def test_absence_leaves_elsewhere_empty_when_dates_match(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Последняя активность человека — здесь же: колонка пустая, страница рисует прочерк."""
    location, _platform_obj, user = _absence_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/absence",
        params={"min_runs": 3, "min_missed": 1},
    )
    row = next(item for item in response.json()["items"] if item["name"] == "Пропавший Иванов")
    assert row["elsewhere_date_display"] is None
    assert row["elsewhere_hint"] is None


# ===== Свод по пробежке =====


def _svod_fixture(db_session: Session) -> tuple[Location, Event, User]:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    past = _event(db_session, platform, location, date(2026, 2, 28), 1)
    today = _event(db_session, platform, location, date(2026, 3, 7), 2)

    newcomer = _participant(db_session, platform, f"{suffix}-new", "Новичок Иванов")
    _run(db_session, today, newcomer, finish_time_sec=1500, position=2)

    veteran = _participant(db_session, platform, f"{suffix}-vet", "Ветеран Петров")
    _run(db_session, past, veteran, finish_time_sec=1300)
    _run(db_session, today, veteran, finish_time_sec=1200, position=1)

    # Волонтёр с новой ролью: раньше был секундомером, сегодня — маршал.
    volunteer = _participant(db_session, platform, f"{suffix}-vol", "Волонтёр Сидоров")
    _volunteer(db_session, past, volunteer, "Секундомер")
    _volunteer(db_session, today, volunteer, "Маршал")

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор Кузнецов")
    _volunteer(db_session, today, organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    return location, today, user


def test_svod_rows_and_flags(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    location, today, user = _svod_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    dates = client.get(f"/api/organizer/{location.external_key}/event-dates")
    assert dates.status_code == 200
    assert str(today.id) in {item["event_id"] for item in dates.json()["items"]}

    response = client.get(
        f"/api/organizer/{location.external_key}/event-report",
        params={"event_id": str(today.id)},
    )
    assert response.status_code == 200
    payload = response.json()

    runners = {row["name"]: row for row in payload["runners"]}
    assert runners["Новичок Иванов"]["first_in_system"] is True
    assert runners["Новичок Иванов"]["location_runs_count"] == 1
    # Ветеран улучшил время 1300 → 1200: личный рекорд и рекорд локации.
    assert runners["Ветеран Петров"]["is_pb"] is True
    assert runners["Ветеран Петров"]["is_location_pb"] is True
    assert runners["Ветеран Петров"]["location_runs_count"] == 2
    # Первым в протоколе идёт первое место.
    assert payload["runners"][0]["position"] == 1

    volunteers = {row["name"]: row for row in payload["volunteers"]}
    assert volunteers["Волонтёр Сидоров"]["new_roles"] == ["Маршал"]
    assert volunteers["Волонтёр Сидоров"]["first_volunteering"] is False
    assert volunteers["Организатор Кузнецов"]["first_volunteering"] is True


def test_svod_rejects_event_of_another_location(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    location, _today, user = _svod_fixture(db_session)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    other_location = _location(db_session, platform, f"other-{uuid4().int % 1_000_000}")
    other_event = _event(db_session, platform, other_location, date(2026, 3, 7), 1)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/event-report",
        params={"event_id": str(other_event.id)},
    )
    assert response.status_code == 404


def test_svod_xlsx_export(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    location, today, user = _svod_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/event-report.xlsx",
        params={"event_id": str(today.id)},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=" in response.headers["content-disposition"]

    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames[:2] == ["Бегуны", "Волонтёры"]
    runners_sheet = workbook["Бегуны"]
    names = {row[1] for row in runners_sheet.iter_rows(min_row=2, values_only=True)}
    assert "Ветеран Петров" in names


# ===== Новые инструменты: пост, юбилеи, новички, скамейка =====


def test_event_post_uses_site_signature(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    location, today, user = _svod_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/event-post",
        params={"event_id": str(today.id)},
    )
    assert response.status_code == 200
    post = response.json()["post_text"]
    # Подпись кабинета ведёт на сайт, а не на канал (решение Дмитрия 16.08.2026).
    assert "run5k.run" in post
    assert "t.me/popov_way" not in post
    assert "Статистика пробежки" in post


def test_milestones_shows_upcoming_jubilee(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Участник с 9 пробежками здесь — в шаге от юбилея 10."""
    from datetime import date as date_type

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    today = date_type.today()
    events = [
        _event(db_session, platform, location, today - timedelta(days=7 * (9 - i)), i + 1)
        for i in range(9)
    ]
    runner = _participant(db_session, platform, f"{suffix}-r9", "Почти Юбиляр")
    for event in events:
        _run(db_session, event, runner)

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор")
    _volunteer(db_session, events[-1], organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(f"/api/organizer/{location.external_key}/milestones")
    assert response.status_code == 200
    items = response.json()["items"]
    runs_here = [
        item for item in items if item["name"] == "Почти Юбиляр" and item["kind"] == "runs_here"
    ]
    assert len(runs_here) == 1
    assert runs_here[0]["current"] == 9
    assert runs_here[0]["milestone"] == 10
    assert runs_here[0]["remaining"] == 1


def test_newcomers_retention(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Дебютант, вернувшийся сюда, — удержан; гость с другой локации — не новичок."""
    from datetime import date as date_type

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    other = _location(db_session, platform, f"org-other-{suffix}")
    today = date_type.today()
    e1 = _event(db_session, platform, location, today - timedelta(days=21), 1)
    e2 = _event(db_session, platform, location, today - timedelta(days=14), 2)
    e3 = _event(db_session, platform, location, today - timedelta(days=7), 3)
    other_event = _event(db_session, platform, other, today - timedelta(days=60), 1)

    returned = _participant(db_session, platform, f"{suffix}-ret", "Вернувшийся Новичок")
    _run(db_session, e1, returned)
    _run(db_session, e2, returned)

    lost = _participant(db_session, platform, f"{suffix}-lost", "Потерянный Новичок")
    _run(db_session, e1, lost)

    # Гость: дебютировал на другой локации, сюда пришёл уже бывалым.
    guest = _participant(db_session, platform, f"{suffix}-guest", "Гость Бывалый")
    _run(db_session, other_event, guest)
    _run(db_session, e2, guest)

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор")
    _volunteer(db_session, e3, organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/newcomers", params={"days": 90}
    )
    assert response.status_code == 200
    payload = response.json()
    names = {item["name"] for item in payload["items"]}
    assert names == {"Вернувшийся Новичок", "Потерянный Новичок"}
    by_name = {item["name"]: item for item in payload["items"]}
    assert by_name["Вернувшийся Новичок"]["returned_here"] is True
    assert by_name["Потерянный Новичок"]["returned_here"] is False
    # Оба дебютировали до последнего события → оба в знаменателе, вернулся один.
    assert payload["eligible_total"] == 2
    assert payload["retention_pct"] == 50


def test_volunteer_bench_candidates_and_statuses(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Скамейка: наверху те, кого звать, а не ветераны с максимумом волонтёрств."""
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    events = [
        _event(db_session, platform, location, date(2026, 1, 3) + timedelta(days=7 * i), i + 1)
        for i in range(14)
    ]

    # Кандидат №1: бегает 12 раз, волонтёрств здесь нет вообще.
    runner = _participant(db_session, platform, f"{suffix}-run", "Бегун Незваный")
    for event in events[:12]:
        _run(db_session, event, runner)

    # Выпал из команды: волонтёрил в самом начале, но продолжает бегать.
    lapsed = _participant(db_session, platform, f"{suffix}-lap", "Волонтёр Выпавший")
    _volunteer(db_session, events[0], lapsed, "Маршал")
    for event in events[6:]:
        _run(db_session, event, lapsed)

    # В строю: волонтёрил на последнем событии.
    active = _participant(db_session, platform, f"{suffix}-act", "Волонтёр Действующий")
    for event in events[:3]:
        _volunteer(db_session, event, active, "Секундомер")
    _volunteer(db_session, events[-1], active, "Секундомер")

    # Случайный бегун ниже порога — в выдачу не попадает.
    rare = _participant(db_session, platform, f"{suffix}-rare", "Бегун Разовый")
    _run(db_session, events[0], rare)

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор")
    _volunteer(db_session, events[-1], organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(f"/api/organizer/{location.external_key}/volunteers")
    assert response.status_code == 200
    payload = response.json()
    by_name = {item["name"]: item for item in payload["items"]}

    # Бегун без волонтёрств попал в выдачу и помечен кандидатом.
    assert by_name["Бегун Незваный"]["status"] == "never"
    assert by_name["Бегун Незваный"]["is_candidate"] is True
    assert by_name["Бегун Незваный"]["runs_here"] == 12
    assert by_name["Бегун Незваный"]["last_vol_display"] is None

    # Выпавший: волонтёрил давно, но бегает — тоже кандидат.
    assert by_name["Волонтёр Выпавший"]["status"] == "paused"
    assert by_name["Волонтёр Выпавший"]["is_candidate"] is True
    assert by_name["Волонтёр Выпавший"]["runs_after_last_vol"] == 8

    # Действующий волонтёр кандидатом не считается.
    assert by_name["Волонтёр Действующий"]["status"] == "active"
    assert by_name["Волонтёр Действующий"]["is_candidate"] is False

    # Ниже порога пробежек — в списке нет.
    assert "Бегун Разовый" not in by_name

    # Порядок: кандидаты первыми, действующие — ниже.
    order = [item["name"] for item in payload["items"]]
    assert order.index("Бегун Незваный") < order.index("Волонтёр Действующий")
    assert payload["candidates_total"] >= 2


def test_event_post_templates(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Каждый шаблон собирается и несёт свой контент; неизвестный — 404."""
    location, today, user = _svod_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    def _post(template: str):
        return client.get(
            f"/api/organizer/{location.external_key}/event-post",
            params={"event_id": str(today.id), "template": template},
        )

    volunteers = _post("volunteers")
    assert volunteers.status_code == 200
    text = volunteers.json()["post_text"]
    assert "Команда волонтёров" in text
    # Формат Мещерского: эмодзи роли, роль — имена; директор первым.
    assert "🦺 Маршал — Волонтёр Сидоров" in text
    lines = text.splitlines()
    first_role_line = next(line for line in lines if line.startswith("1."))
    assert "Директор забега" in first_role_line

    newcomers = _post("newcomers")
    assert newcomers.status_code == 200
    text = newcomers.json()["post_text"]
    assert "Новичок Иванов" in text
    assert "Первый финиш" in text

    stats = _post("stats")
    assert stats.status_code == 200
    text = stats.json()["post_text"]
    assert "Личные рекорды обновили" in text
    assert "Ветеран Петров" in text

    milestones = _post("milestones")
    assert milestones.status_code == 200

    assert _post("unknown").status_code == 404


def test_upcoming_post_needs_no_event(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """«Юбилеи завтра» строится по локации; человек с 9 пробежками — в посте."""
    from datetime import date as date_type

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    today = date_type.today()
    events = [
        _event(db_session, platform, location, today - timedelta(days=7 * (9 - i)), i + 1)
        for i in range(9)
    ]
    runner = _participant(db_session, platform, f"{suffix}-r9", "Почти Юбиляр")
    for event in events:
        _run(db_session, event, runner)
    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор")
    _volunteer(db_session, events[-1], organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{location.external_key}/event-post",
        params={"template": "upcoming"},
    )
    assert response.status_code == 200
    text = response.json()["post_text"]
    assert "Почти Юбиляр" in text
    assert "10-я" in text

    # Обычному шаблону событие обязательно.
    missing = client.get(
        f"/api/organizer/{location.external_key}/event-post",
        params={"template": "full"},
    )
    assert missing.status_code == 422


def test_travelers_post(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Постоянный участник на чужом старте в ту же дату попадает в «Наши в гостях»."""
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    home = _location(db_session, platform, f"org-home-{suffix}")
    away = _location(db_session, platform, f"org-away-{suffix}")
    events = [
        _event(db_session, platform, home, date(2026, 1, 3) + timedelta(days=7 * i), i + 1)
        for i in range(6)
    ]
    away_event = _event(db_session, platform, away, events[-1].event_date, 1)

    # Свой: 5 финишей дома, на шестую субботу уехал в чужой парк.
    tourist = _participant(db_session, platform, f"{suffix}-tour", "Турист Домашний")
    for event in events[:5]:
        _run(db_session, event, tourist)
    _run(db_session, away_event, tourist, finish_time_sec=1500)

    # Случайный гость: один финиш дома — «своим» не считается.
    random_guest = _participant(db_session, platform, f"{suffix}-rnd", "Случайный Гость")
    _run(db_session, events[0], random_guest)
    _run(db_session, away_event, random_guest)

    # Постоянный гость: 5 финишей у нас, но дома (по трём ступеням) он в другом
    # парке — 6 финишей там. В рубрику попадать не должен.
    away_events = [
        _event(db_session, platform, away, date(2025, 10, 4) + timedelta(days=7 * i), 10 + i)
        for i in range(5)
    ]
    frequent_guest = _participant(db_session, platform, f"{suffix}-fg", "Постоянный Гость")
    for event in events[:5]:
        _run(db_session, event, frequent_guest)
    for event in away_events:
        _run(db_session, event, frequent_guest)
    _run(db_session, away_event, frequent_guest)

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор")
    _volunteer(db_session, events[-1], organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    response = client.get(
        f"/api/organizer/{home.external_key}/event-post",
        params={"event_id": str(events[-1].id), "template": "travelers"},
    )
    assert response.status_code == 200
    text = response.json()["post_text"]
    assert "Наши в гостях" in text
    assert "Турист Домашний" in text
    # Рядом с именем — сколько пробежек у нас на дату события (5 из 6 стартов).
    assert "Турист Домашний (5 пробежек у нас)" in text
    assert f"Локация org-away-{suffix}" in text
    assert "Случайный Гость" not in text
    # ≥5 финишей у нас есть, но домашняя локация — чужой парк.
    assert "Постоянный Гость" not in text


def test_parse_volunteer_roster() -> None:
    """Парсер таблицы записи 5 вёрст: даты, роли, занятые клетки."""
    from app.services.organizer_roster_service import parse_volunteer_roster

    html = """
    <table class="resultsTable"><tbody>
      <tr><th><strong>Роль</strong></th><th><strong>22.08.2026</strong></th><th><strong>29.08.2026</strong></th></tr>
      <tr><td><strong>Организатор</strong></td><td>Сергей МИСЮКОВ</td><td></td></tr>
      <tr><td><strong>Секундомер</strong></td><td></td><td>Надежда ВОРОБЬЕВА</td></tr>
    </tbody></table>
    """
    payload = parse_volunteer_roster(html, "https://example.test/volunteer/")
    assert payload is not None
    assert payload["dates"] == ["22.08.2026", "29.08.2026"]
    roles = {row["role"]: row["filled"] for row in payload["roles"]}
    assert roles["Организатор"] == {"22.08.2026": "Сергей МИСЮКОВ"}
    assert roles["Секундомер"] == {"29.08.2026": "Надежда ВОРОБЬЕВА"}


def test_vacancies_post_from_roster(
    client: TestClient,
    db_session: Session,
    fake_redis: fakeredis.FakeRedis,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """«Нужны волонтёры» строится из живой записи: пустая клетка = вакансия."""
    location, _today, user = _svod_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    monkeypatch.setattr(
        "app.services.organizer_roster_service.fetch_volunteer_roster",
        lambda slug, **kwargs: {
            "source_url": f"https://5verst.ru/{slug}/volunteer/",
            "dates": ["22.08.2026"],
            "roles": [
                {"role": "Организатор", "filled": {"22.08.2026": "Сергей МИСЮКОВ"}},
                {"role": "Секундомер", "filled": {}},
                {"role": "Замыкающий", "filled": {}},
            ],
        },
    )

    response = client.get(
        f"/api/organizer/{location.external_key}/event-post",
        params={"template": "vacancies"},
    )
    assert response.status_code == 200
    text = response.json()["post_text"]
    assert "Нужны волонтёры на 22.08.2026" in text
    assert "❗️ ⏱️ Секундомер" in text
    assert "❗️ 🐢 Замыкающий" in text
    assert "✅ 🪇 Организатор — Сергей МИСЮКОВ" in text
    assert "volunteer/" in text


def test_vacancies_post_fallback_without_roster(
    client: TestClient,
    db_session: Session,
    fake_redis: fakeredis.FakeRedis,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Без живой записи — запасной список ключевых позиций."""
    location, _today, user = _svod_fixture(db_session)
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    monkeypatch.setattr(
        "app.services.organizer_roster_service.fetch_volunteer_roster",
        lambda slug, **kwargs: None,
    )

    response = client.get(
        f"/api/organizer/{location.external_key}/event-post",
        params={"template": "vacancies"},
    )
    assert response.status_code == 200
    text = response.json()["post_text"]
    # Ключевые позиции есть, факультативных (например, «Видеограф») нет.
    assert "Секундомер" in text
    assert "Директор забега" in text
    assert "Видеограф" not in text
    assert "оставьте только роли" in text


def test_analytics_endpoints(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Команда/посещаемость/портрет/бенчмарк: базовые расчёты на одной фикстуре."""
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-loc-{suffix}")
    from datetime import date as date_type

    today = date_type.today()
    events = [
        _event(db_session, platform, location, today - timedelta(days=7 * (6 - i)), i + 1)
        for i in range(6)
    ]

    # Секундомер все 6 стартов — bus-фактор 1, «держится на одном».
    timer = _participant(db_session, platform, f"{suffix}-tm", "Вечный Секундомер")
    for event in events:
        _volunteer(db_session, event, timer, "Секундомер")
    # Маршалы разные — ротация 100%.
    for index, event in enumerate(events[:3]):
        marshal = _participant(db_session, platform, f"{suffix}-m{index}", f"Маршал {index}")
        _volunteer(db_session, event, marshal, "Маршал")

    # Бегуны: мужчина и женщина с возрастными группами и клубом.
    man = _participant(db_session, platform, f"{suffix}-man", "Бегун Клубный")
    man.gender = "male"
    man.club_name = "ТестКлуб"
    woman = _participant(db_session, platform, f"{suffix}-w", "Бегунья Быстрая")
    woman.gender = "female"
    for event in events:
        db_session.add(
            RunResult(
                event_id=event.id,
                participant_id=man.id,
                external_result_key=f"run-{uuid4()}",
                finish_time_sec=1200,
                finish_time_display="00:20:00",
                age_category="М35-39 (1)",
                status="finished",
            )
        )
        db_session.add(
            RunResult(
                event_id=event.id,
                participant_id=woman.id,
                external_result_key=f"run-{uuid4()}",
                finish_time_sec=1400,
                finish_time_display="00:23:20",
                age_category="Ж30-34 (1)",
                status="finished",
            )
        )

    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор")
    _volunteer(db_session, events[-1], organizer, "Организатор")
    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "organizer")

    # --- Команда: секундомер — узкое место, маршалы ротируются.
    team = client.get(f"/api/organizer/{location.external_key}/team")
    assert team.status_code == 200
    roles = {r["role"]: r for r in team.json()["roles"]}
    assert roles["Секундомер"]["bus_factor"] == 1
    assert roles["Секундомер"]["top_name"] == "Вечный Секундомер"
    assert roles["Секундомер"]["is_critical"] is True
    assert roles["Маршал"]["people"] == 3
    assert roles["Маршал"]["rotation_pct"] == 100
    # Ключевые роли с низким bus-фактором — первыми (Директор и Секундомер оба bus=1).
    first = team.json()["roles"][0]
    assert first["is_critical"] is True and first["bus_factor"] == 1

    # --- Посещаемость: 6 стартов по 2 финишёра.
    attendance = client.get(f"/api/organizer/{location.external_key}/attendance")
    assert attendance.status_code == 200
    payload = attendance.json()
    assert payload["events_total"] == 6
    assert payload["record_finishers"] == 2
    assert all(month["avg_finishers"] == 2.0 for month in payload["months"])

    # --- Портрет: пол пополам, группы чистятся от «(1)», клуб виден.
    audience = client.get(f"/api/organizer/{location.external_key}/audience")
    assert audience.status_code == 200
    payload = audience.json()
    assert payload["people_total"] == 2
    groups = {g["group"] for g in payload["age_groups"]}
    assert groups == {"М35-39", "Ж30-34"}
    genders = {g["label"]: g["share_pct"] for g in payload["genders"]}
    assert genders == {"Мужчины": 50.0, "Женщины": 50.0}
    assert payload["clubs"][0]["club"] == "ТестКлуб"

    # --- Бенчмарк: локация в выборке и отвечает без ошибок.
    benchmark = client.get(
        f"/api/organizer/{location.external_key}/benchmark", params={"scope": "network"}
    )
    assert benchmark.status_code == 200
    payload = benchmark.json()
    if payload["metrics"]:
        finishers = next(m for m in payload["metrics"] if m["key"] == "avg_finishers")
        assert finishers["our_value"] == 2.0


# ===== Расписание стартов: парсер schedule_text =====


def test_schedule_parser() -> None:
    from datetime import date as date_type

    from app.services.location_schedule_service import parse_schedule_text, start_time_for_date

    # Обычный случай: «каждую субботу в 9:00».
    plain = parse_schedule_text("Старт проходит по адресу: Город, улица. Старт каждую субботу в 9:00.")
    assert plain == [{"from_month": 1, "to_month": 12, "time": "09:00"}]

    # Сезонное расписание с переломом через декабрь.
    seasonal = parse_schedule_text(
        "Старт каждую субботу в 9:00 (с сентября по май), в 8:00 (июнь — август)."
    )
    assert len(seasonal) == 2
    assert start_time_for_date(seasonal, date_type(2026, 1, 10)).hour == 9
    assert start_time_for_date(seasonal, date_type(2026, 7, 11)).hour == 8

    # Не-утреннее одинокое число временем старта не считается.
    assert parse_schedule_text("Парк работает до 22:00 ежедневно.") == []


# ===== Протоколы: задержка выгрузки и журнал правок =====


def test_protocols_endpoint_delay_formula(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """Задержка = появление − (старт + финиш последнего), с поясом локации."""
    from datetime import datetime as datetime_type
    from datetime import time as time_type
    from datetime import timezone as timezone_type

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-prot-{suffix}")
    # Локация на два часа восточнее Москвы: старт 09:00 местного = 04:00 UTC.
    location.tz_offset_moscow = 2
    db_session.add(
        LocationDescription(
            location_id=location.id,
            schedule_text="Старт каждую субботу в 9:00.",
            schedule_parsed=[{"from_month": 1, "to_month": 12, "time": "09:00"}],
        )
    )

    event_day = date.today() - timedelta(days=7)
    event = _event(db_session, platform, location, event_day, 1)
    runner = _participant(db_session, platform, f"{suffix}-r", "Бегун Последний")
    # Последний финиш через 40 минут после старта.
    _run(db_session, event, runner, finish_time_sec=2400, position=1)

    organizer = _participant(db_session, platform, f"{suffix}-org", "Директор Дня")
    _volunteer(db_session, event, organizer, "Организатор")

    # Протокол замечен через 2 часа после финиша последнего:
    # старт 04:00 UTC + 40 мин = 04:40, факт в 06:40 UTC.
    first_seen = datetime_type.combine(event_day, time_type(6, 40), tzinfo=timezone_type.utc)
    db_session.add(
        ProtocolUploadFact(
            location_id=location.id,
            event_date=event_day,
            first_seen_at=first_seen,
            source="site",
        )
    )
    db_session.add(
        ProtocolRevision(
            event_id=event.id,
            detected_at=first_seen + timedelta(hours=5),
            kind="times_changed",
            details={"time_changes_total": 2},
        )
    )

    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "protocol-organizer")

    response = client.get(f"/api/organizer/{location.external_key}/protocols")
    assert response.status_code == 200
    payload = response.json()
    assert payload["supported"] is True
    assert payload["tz_offset_moscow"] == 2

    item = next(row for row in payload["items"] if row["date"] == event_day.isoformat())
    assert item["start_time"] == "09:00"
    assert item["last_finish_display"] == "40:00"
    assert item["delay_hours"] == 2.0
    assert item["level"] == "green"
    assert item["directors"] == ["Директор Дня"]
    assert len(item["revisions"]) == 1
    assert item["revisions"][0]["kind"] == "times_changed"


def test_protocol_revision_ignores_identified_unknown(db_session: Session) -> None:
    """«Неизвестный стал известным» — не правка; сдвиг времени — правка."""
    from app.sync.five_verst_protocol import record_protocol_revision

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-rev-{suffix}")
    event = _event(db_session, platform, location, date.today() - timedelta(days=7), 1)
    db_session.commit()

    # Только «неизвестный стал известным» (та же позиция и время): журнал пуст.
    record_protocol_revision(
        db_session,
        event.id,
        {"k1": (1, 1200, "finished"), "k2": (2, 1300, "unknown")},
        {"k1": (1, 1200, "finished"), "k2-known": (2, 1300, "finished")},
    )
    assert db_session.query(ProtocolRevision).filter(ProtocolRevision.event_id == event.id).count() == 0

    # Время изменилось — правка фиксируется.
    record_protocol_revision(
        db_session,
        event.id,
        {"k1": (1, 1200, "finished")},
        {"k1": (1, 1180, "finished")},
    )
    revisions = db_session.query(ProtocolRevision).filter(ProtocolRevision.event_id == event.id).all()
    assert len(revisions) == 1
    assert revisions[0].kind == "times_changed"
    assert revisions[0].details["time_changes_total"] == 1


def test_protocol_revision_catches_removal_next_to_identified_unknown(
    db_session: Session,
) -> None:
    """Пропажа известной строки не прячется за парой «неизвестный → имя».

    Удаление не даёт добавления, поэтому старое сравнение отфильтрованных
    наборов (только unknown-удаления против known-добавлений) её проглатывало.
    """
    from app.sync.five_verst_protocol import record_protocol_revision

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-rev2-{suffix}")
    event = _event(db_session, platform, location, date.today() - timedelta(days=7), 1)
    db_session.commit()

    record_protocol_revision(
        db_session,
        event.id,
        {
            "k1": (1, 1200, "finished"),
            "k2": (2, 1300, "unknown"),
            "k3": (3, 1400, "finished"),  # последний известный — пропал
        },
        {
            "k1": (1, 1200, "finished"),
            "k2-known": (2, 1300, "finished"),
        },
    )
    revisions = db_session.query(ProtocolRevision).filter(ProtocolRevision.event_id == event.id).all()
    assert len(revisions) == 1
    assert revisions[0].kind == "results_changed"
    assert revisions[0].details["removed"] == 2
    assert revisions[0].details["added"] == 1


def test_health_endpoint_smoke(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-health-{suffix}")
    event = _event(db_session, platform, location, date.today() - timedelta(days=7), 1)
    organizer = _participant(db_session, platform, f"{suffix}-org", "Организатор Здоровья")
    _volunteer(db_session, event, organizer, "Организатор")
    photographer = _participant(db_session, platform, f"{suffix}-ph", "Фотограф Тестовый")
    _volunteer(db_session, event, photographer, "Фотограф")

    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "health-organizer")

    response = client.get(f"/api/organizer/{location.external_key}/health")
    assert response.status_code == 200
    payload = response.json()
    keys = [indicator["key"] for indicator in payload["indicators"]]
    # Фотограф из светофора убран (23.08.2026), _photographer_share остался на будущее.
    # Ротация организаторов стоит сразу за ротацией волонтёров: та же по смыслу
    # роль, но выгорание в ней закрывает площадку целиком (Дмитрий 03.09.2026).
    assert keys == [
        "protocol",
        "rotation",
        "director_rotation",
        "attendance",
        "protocol_quality",
        "newcomers",
    ]
    # Подсказка «что это» обязательна у всех; advice остался у протокола
    # («на что влияет») и у ротации с новичками («как улучшить»).
    for indicator in payload["indicators"]:
        assert indicator["hint"]
    by_key = {indicator["key"]: indicator for indicator in payload["indicators"]}
    assert by_key["protocol"]["advice"].startswith("На что влияет")
    assert by_key["attendance"]["advice"] is None
    assert by_key["protocol_quality"]["advice"] is None


def test_protocol_watch_records_first_seen_once(
    db_session: Session, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Наблюдатель пишет факт один раз: повторный прогон не двигает first_seen_at."""
    from app.platform_adapters.canonical import CanonicalEventSummary
    from app.sync import five_verst_protocol_watch as watch

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    location = _location(db_session, platform, f"org-watch-{suffix}")
    db_session.commit()

    event_day = date.today() - timedelta(days=1)
    summary = CanonicalEventSummary(
        external_event_key=f"{location.external_key}-{event_day.isoformat()}",
        event_date=event_day,
        event_number=10,
        location_external_key=location.external_key,
        location_name=location.name,
    )
    test_summary = CanonicalEventSummary(
        external_event_key=f"{location.external_key}-test",
        event_date=event_day,
        event_number=11,
        location_external_key=location.external_key,
        location_name=location.name,
        is_test_event=True,
    )
    monkeypatch.setattr(
        watch.bulk_parser, "fetch_latest_results", lambda: ([summary, test_summary], "<html>")
    )

    first = watch.record_protocol_upload_facts(db_session)
    assert first.new_facts == 1
    fact = (
        db_session.query(ProtocolUploadFact)
        .filter(ProtocolUploadFact.location_id == location.id)
        .one()
    )
    seen_at = fact.first_seen_at

    second = watch.record_protocol_upload_facts(db_session)
    assert second.new_facts == 0
    db_session.expire_all()
    fact_again = (
        db_session.query(ProtocolUploadFact)
        .filter(ProtocolUploadFact.location_id == location.id)
        .one()
    )
    assert fact_again.first_seen_at == seen_at


def test_protocol_watch_marks_cold_start_as_unconfirmed(
    db_session: Session, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Первый прогон видит протоколы уже лежащими — момент появления неизвестен.

    02.09.2026 наблюдатель на своём первом прогоне записал 20 площадок разом,
    как будто все протоколы появились в 21:00. У Ставрополя протокол лежал с
    29 августа, а кабинет обвинял организатора в задержке 109 часов.
    """
    from app.platform_adapters.canonical import CanonicalEventSummary
    from app.sync import five_verst_protocol_watch as watch

    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "five_verst", "5 вёрст")
    cold = _location(db_session, platform, f"org-cold-{suffix}")
    warm = _location(db_session, platform, f"org-warm-{suffix}")
    db_session.commit()

    event_day = date.today() - timedelta(days=1)

    def _summary(location: Location) -> CanonicalEventSummary:
        return CanonicalEventSummary(
            external_event_key=f"{location.external_key}-{event_day.isoformat()}",
            event_date=event_day,
            event_number=10,
            location_external_key=location.external_key,
            location_name=location.name,
        )

    # Первый прогон: наблюдения до него не было.
    monkeypatch.setattr(watch.bulk_parser, "fetch_latest_results", lambda: ([_summary(cold)], "<html>"))
    first = watch.record_protocol_upload_facts(db_session)
    assert first.new_facts == 1
    assert first.unconfirmed_facts == 1
    cold_fact = (
        db_session.query(ProtocolUploadFact)
        .filter(ProtocolUploadFact.location_id == cold.id)
        .one()
    )
    assert cold_fact.first_seen_confirmed is False

    # Следующий прогон идёт сразу за предыдущим — наблюдение непрерывно, и
    # протокол, которого минуту назад не было, записывается подтверждённым.
    monkeypatch.setattr(watch.bulk_parser, "fetch_latest_results", lambda: ([_summary(warm)], "<html>"))
    second = watch.record_protocol_upload_facts(db_session)
    assert second.new_facts == 1
    assert second.unconfirmed_facts == 0
    warm_fact = (
        db_session.query(ProtocolUploadFact)
        .filter(ProtocolUploadFact.location_id == warm.id)
        .one()
    )
    assert warm_fact.first_seen_confirmed is True


def test_health_skips_protocol_for_non_five_verst(
    client: TestClient, db_session: Session, fake_redis: fakeredis.FakeRedis
) -> None:
    """У локации без 5в-половины индикатора скорости протокола в светофоре нет."""
    suffix = str(uuid4().int % 1_000_000)
    platform = _platform(db_session, "s95", "с95")
    location = _location(db_session, platform, f"org-s95-{suffix}")
    event = _event(db_session, platform, location, date.today() - timedelta(days=7), 1)
    organizer = _participant(db_session, platform, f"{suffix}-org", "Директор С95")
    _volunteer(db_session, event, organizer, "Директор 25")

    user = User(telegram_id=int(uuid4().int % 10_000_000_000), telegram_username=f"u{suffix}")
    db_session.add(user)
    db_session.flush()
    _link(db_session, user, platform, organizer)
    db_session.commit()
    fake_redis.delete(LOCATIONS_INDEX_CACHE_KEY)
    _login(client, user.telegram_id or 0, "s95-organizer")

    response = client.get(f"/api/organizer/{location.external_key}/health")
    assert response.status_code == 200
    keys = [indicator["key"] for indicator in response.json()["indicators"]]
    assert "protocol" not in keys
    assert keys[0] == "rotation"

    protocols = client.get(f"/api/organizer/{location.external_key}/protocols")
    assert protocols.status_code == 200
    assert protocols.json()["supported"] is False

#!/usr/bin/env python3
"""Import location catalog from reviewed XLSX into DB and JSON seed."""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_ROOT = SCRIPT_DIR.parent
REPO_ROOT = BACKEND_ROOT.parent

if (REPO_ROOT / "data").is_dir():
    DATA_ROOT = REPO_ROOT
elif Path("/data").is_dir():
    DATA_ROOT = Path("/")
else:
    DATA_ROOT = REPO_ROOT

DEFAULT_XLSX = DATA_ROOT / "data" / "location_mapping_draft.xlsx"
DEFAULT_JSON = DATA_ROOT / "data" / "location_catalog.json"

PARKRUN_SLUG_ALIASES: dict[str, str] = {
    "gorky-park": "gorkypark",
    "park-850-letiya-moskvy": "park850letiyamoskvy",
    "tula-central": "tulacentral",
    "tver-rechnoy-vokzal": "tverrechnoyvokzal",
    "moskovsky-park-pobedy": "moskovskyparkpobedy",
    "lobnya-gorodskoy-park": "lobnyagorodskoypark",
    "serpukhov-gorodskoy-bor": "serpukhovgorodskoybor",
    "gatchina-prioratsky": "gatchinaprioratsky",
    "balashikha-zarechnaya": "balashikhazarechnaya",
    "babushkinsky-na-yauze": "babushkinskynayauze",
    "pokrovskoe-streshnevo": "pokrovskoestreshnevo",
    "severnoe-tushino": "severnoetushino",
    "olimpiyskaya-derevnya": "olimpiyskayaderevnya",
    "lesopark-severny": "lesoparkseverny",
}


PARKRUN_TO_S95_SLUG: dict[str, str] = {
    "izmailovo": "izmailovo",
    "tsaritsyno": "tsaritsyno",
    "olimpiyskaya-derevnya": "olimpiyskaya_derevnya",
    "tver-rechnoy-vokzal": "parkzhrun",
    "gatchina-prioratsky": "gatchina",
}


def five_verst_external_key(parkrun_slug: str, actual_slug: str | None) -> str:
    if actual_slug and actual_slug in PARKRUN_SLUG_ALIASES.values():
        return actual_slug
    if actual_slug and "-" not in actual_slug:
        return actual_slug
    return PARKRUN_SLUG_ALIASES.get(parkrun_slug, norm_slug(parkrun_slug))


def s95_external_key(parkrun_slug: str, actual_slug: str | None) -> str:
    if actual_slug and "_" in str(actual_slug):
        return str(actual_slug)
    return PARKRUN_TO_S95_SLUG.get(parkrun_slug, actual_slug and str(actual_slug) or norm_slug(parkrun_slug))

URL_5VERST = re.compile(r"5verst\.ru/([^/\"'\s]+)", re.I)
URL_S95 = re.compile(r"s95\.ru/events/([^/\"'\s]+)", re.I)


@dataclass
class CatalogEntry:
    parkrun_location: str
    canonical_name: str
    active_platform: str | None
    legacy_parkrun_slug: str
    is_closed: bool = False
    notes: str = ""
    links: list[dict[str, str]] = field(default_factory=list)


def norm_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def parse_url_links(text: str) -> list[tuple[str, str]]:
    links: list[tuple[str, str]] = []
    for match in URL_5VERST.finditer(text or ""):
        links.append(("five_verst", match.group(1)))
    for match in URL_S95.finditer(text or ""):
        links.append(("s95", match.group(1)))
    return links


def parse_system(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip().lower().replace("ё", "е")
    if "5" in text and "верст" in text:
        return "five_verst"
    if "с95" in text or "s95" in text:
        return "s95"
    if "runpark" in text:
        return "runpark"
    if "parkrun" in text:
        return "parkrun"
    return None


def is_closed_confirmed(confirmed: str | None, comment: str | None) -> bool:
    blob = f"{confirmed or ''} {comment or ''}".lower()
    return "более не существ" in blob or "не существ" in blob


def is_runpark_only(confirmed: str | None, system: str | None) -> bool:
    blob = (confirmed or "").lower()
    return "runpark" in blob or system == "нет (только parkrun)"


def row_to_entry(row: tuple) -> CatalogEntry | None:
    parkrun_name, current_loc, system, comment, confirmed, pr_slug, actual_slug, _confidence = row[:8]
    if not pr_slug:
        return None

    confirmed_text = (confirmed or "").strip()
    confirmed_lower = confirmed_text.lower()
    closed = is_closed_confirmed(confirmed_text, comment)
    runpark = is_runpark_only(confirmed_text, system)
    active = parse_system(system)

    url_links = parse_url_links(confirmed_text)
    if url_links:
        active = url_links[0][0]

    if closed:
        return CatalogEntry(
            parkrun_location=str(parkrun_name or ""),
            canonical_name=str(current_loc or parkrun_name or pr_slug),
            active_platform=None,
            legacy_parkrun_slug=str(pr_slug),
            is_closed=True,
            notes="; ".join(filter(None, [comment, confirmed_text])),
            links=[{"platform": "parkrun", "external_key": str(pr_slug)}],
        )

    if runpark:
        return CatalogEntry(
            parkrun_location=str(parkrun_name or ""),
            canonical_name=str(current_loc or parkrun_name or pr_slug),
            active_platform="runpark",
            legacy_parkrun_slug=str(pr_slug),
            notes="; ".join(filter(None, [comment, confirmed_text])),
            links=[{"platform": "parkrun", "external_key": str(pr_slug)}],
        )

    if url_links:
        canonical = str(current_loc or parkrun_name or "")
        for _, slug in url_links:
            if active == "five_verst" and not canonical:
                canonical = slug
        links = [{"platform": "parkrun", "external_key": str(pr_slug)}]
        for platform, slug in url_links:
            links.append({"platform": platform, "external_key": slug})
        return CatalogEntry(
            parkrun_location=str(parkrun_name or ""),
            canonical_name=canonical or str(parkrun_name or ""),
            active_platform=active,
            legacy_parkrun_slug=str(pr_slug),
            notes="; ".join(filter(None, [comment, confirmed_text])),
            links=links,
        )

    if confirmed_lower in ("да", "yes") or confirmed_lower.startswith("да"):
        pass
    else:
        return None

    if not current_loc or not active:
        return None

    links: list[dict[str, str]] = [{"platform": "parkrun", "external_key": str(pr_slug)}]
    if active == "five_verst":
        links.append({"platform": "five_verst", "external_key": five_verst_external_key(str(pr_slug), actual_slug and str(actual_slug))})
    elif active == "s95":
        links.append({"platform": "s95", "external_key": s95_external_key(str(pr_slug), actual_slug and str(actual_slug))})

    return CatalogEntry(
        parkrun_location=str(parkrun_name or ""),
        canonical_name=str(current_loc),
        active_platform=active,
        legacy_parkrun_slug=str(pr_slug),
        notes=str(comment or ""),
        links=links,
    )


def load_xlsx(path: Path) -> list[CatalogEntry]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["соответствия"]
    entries: list[CatalogEntry] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = row_to_entry(row)
        if entry is not None:
            entries.append(entry)
    wb.close()
    return entries


def export_json(entries: list[CatalogEntry], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [asdict(entry) for entry in entries]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_json(path: Path) -> list[CatalogEntry]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    return [CatalogEntry(**item) for item in raw]


class StaleImportError(RuntimeError):
    """Raised when replace=True would shrink the catalog — likely a stale JSON."""


def import_to_db(
    entries: list[CatalogEntry], *, replace: bool = True, force: bool = False
) -> dict[str, int]:
    backend_path = str(BACKEND_ROOT)
    if backend_path not in sys.path and (BACKEND_ROOT / "app").is_dir():
        sys.path.insert(0, backend_path)

    from app.db.session import get_session_factory
    from app.models import Location, LocationCatalog, LocationCatalogLink, Platform
    from app.services.location_catalog_service import normalize_platform_code

    db = get_session_factory()()
    stats = {"catalog": 0, "links": 0, "linked_locations": 0}
    try:
        platforms = {p.code: p for p in db.query(Platform).all()}
        if replace:
            # replace=True DELETEs the whole catalog before rebuilding from `entries` —
            # if `entries` came from a stale data/location_catalog.json (an old
            # worktree that never pulled later catalog-linking commits), this would
            # silently wipe manually-curated links with no trace in git history.
            # 20.07.2026: found 16/20 local worktrees sitting on a stale copy after a
            # 21-link batch landed in main. Refuse the shrink unless --force.
            current_links = db.query(LocationCatalogLink).count()
            incoming_links = sum(len(e.links) for e in entries)
            if not force and incoming_links < current_links:
                raise StaleImportError(
                    f"Refusing to import: incoming JSON has {incoming_links} links, "
                    f"DB currently has {current_links}. This --import-db would DELETE "
                    "and shrink the catalog — almost always a stale data/location_catalog.json "
                    "(git pull origin main first). If this reduction is intentional "
                    "(e.g. removing a bad link), pass --force."
                )
            db.query(LocationCatalogLink).delete()
            db.query(LocationCatalog).delete()
            db.flush()

        # Отдельные словари для точного и нормализованного slug: если у одной
        # платформы несколько Location с разными реальными external_key нормализуются
        # в одинаковый слаг (легаси-дубли вроде "voronezh-central-park" и
        # "voronezhcentralpark"), точное совпадение не должно затираться более
        # поздним по итерации нормализованным совпадением от другой локации.
        location_by_exact_key: dict[tuple[str, str], Location] = {}
        location_by_norm_key: dict[tuple[str, str], Location] = {}
        for loc, platform in db.query(Location, Platform).join(Platform, Location.platform_id == Platform.id):
            location_by_exact_key[(platform.code, loc.external_key)] = loc
            normalized = norm_slug(loc.external_key)
            if normalized:
                location_by_norm_key.setdefault((platform.code, normalized), loc)

        for entry in entries:
            catalog = LocationCatalog(
                canonical_name=entry.canonical_name,
                legacy_parkrun_slug=entry.legacy_parkrun_slug or None,
                active_platform=entry.active_platform,
                is_closed=entry.is_closed,
                notes=entry.notes or None,
            )
            db.add(catalog)
            db.flush()
            stats["catalog"] += 1

            seen_links: set[tuple[str, str]] = set()
            for link in entry.links:
                platform_code = normalize_platform_code(link["platform"]) or link["platform"]
                external_key = link["external_key"]
                key = (platform_code, external_key)
                if key in seen_links:
                    continue
                seen_links.add(key)

                platform = platforms.get(platform_code)
                if platform is None:
                    continue

                location = location_by_exact_key.get(key) or location_by_norm_key.get(
                    (platform_code, norm_slug(external_key))
                )
                db.add(
                    LocationCatalogLink(
                        catalog_id=catalog.id,
                        platform_id=platform.id,
                        external_key=external_key,
                        location_id=location.id if location else None,
                    )
                )
                stats["links"] += 1
                if location:
                    stats["linked_locations"] += 1

        db.commit()
        return stats
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--import-db", action="store_true", help="Load into PostgreSQL")
    parser.add_argument("--from-json", action="store_true", help="Import from JSON instead of XLSX")
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow --import-db to shrink the catalog (bypasses the stale-JSON guard)",
    )
    args = parser.parse_args()

    if args.from_json:
        entries = load_json(args.json)
    else:
        entries = load_xlsx(args.xlsx)
        export_json(entries, args.json)

    print(f"Entries: {len(entries)}")
    print(f"  closed: {sum(1 for e in entries if e.is_closed)}")
    print(f"  runpark: {sum(1 for e in entries if e.active_platform == 'runpark')}")
    print(f"  five_verst: {sum(1 for e in entries if e.active_platform == 'five_verst')}")
    print(f"  s95: {sum(1 for e in entries if e.active_platform == 's95')}")
    print(f"JSON: {args.json}")

    if args.import_db:
        try:
            stats = import_to_db(entries, force=args.force)
        except StaleImportError as exc:
            print(f"DB import ABORTED: {exc}", file=sys.stderr)
            return 1
        print(f"DB import: {stats}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""Import runpark location correspondence from reviewed XLSX into PostgreSQL.

Usage:
  python scripts/import_runpark_location_mappings.py
  python scripts/import_runpark_location_mappings.py --import-db --geocode
  python scripts/import_runpark_location_mappings.py --from-json --import-db
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = BACKEND_ROOT.parent
DEFAULT_XLSX = REPO_ROOT / "data" / "runpark_import" / "vw_locations_202605271905_matched.xlsx"
DEFAULT_JSON = REPO_ROOT / "data" / "runpark_import" / "runpark_location_mappings.json"


def norm_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def load_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    from app.runpark.mappings import RunparkMappingEntry, parse_xlsx_row

    source_batch = path.stem
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    entries: list[RunparkMappingEntry] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        entries.append(parse_xlsx_row(row, source_batch=source_batch))
    wb.close()
    return [entry.to_dict() for entry in entries]


def load_json(path: Path) -> list[dict]:
    return json.loads(path.read_text(encoding="utf-8"))


def export_json(entries: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(entries, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def geocode_entries(entries: list[dict]) -> list[dict]:
    from app.geo.reverse_geocode import lookup_address
    from app.runpark.mappings import RunparkMappingEntry, fill_missing_geo

    updated: list[dict] = []
    for raw in entries:
        entry = RunparkMappingEntry(**raw)
        needs_geo = not entry.country or not entry.region
        if needs_geo and entry.latitude is not None and entry.longitude is not None:
            geo = lookup_address(entry.latitude, entry.longitude)
            entry = fill_missing_geo(entry, geo)
        updated.append(entry.to_dict())
    return updated


def import_to_db(entries: list[dict], *, replace: bool = True) -> dict[str, int]:
    backend_path = str(BACKEND_ROOT)
    if backend_path not in sys.path and (BACKEND_ROOT / "app").is_dir():
        sys.path.insert(0, backend_path)

    from app.models import (
        Location,
        LocationCatalog,
        LocationCatalogLink,
        Platform,
        RunparkLocationMapping,
    )
    from app.db.session import get_session_factory
    from app.runpark.mappings import RunparkMappingDecision

    db = get_session_factory()()
    stats = {
        "mappings": 0,
        "runpark_locations": 0,
        "catalog_links": 0,
        "matched_locations": 0,
    }
    try:
        platforms = {p.code: p for p in db.query(Platform).all()}
        runpark_platform = platforms.get("runpark")
        if runpark_platform is None:
            raise RuntimeError("runpark platform missing; run alembic upgrade head")

        if replace:
            db.query(RunparkLocationMapping).delete()
            db.flush()

        location_index: dict[tuple[str, str], Location] = {}
        for loc, platform in db.query(Location, Platform).join(Platform, Location.platform_id == Platform.id):
            location_index[(platform.code, loc.external_key)] = loc
            location_index[(platform.code, norm_slug(loc.external_key))] = loc

        catalog_by_slug: dict[str, LocationCatalog] = {}
        for catalog in db.query(LocationCatalog).filter(LocationCatalog.legacy_parkrun_slug.isnot(None)):
            if catalog.legacy_parkrun_slug:
                catalog_by_slug[catalog.legacy_parkrun_slug] = catalog

        for raw in entries:
            matched_location = None
            if raw.get("matched_platform") and raw.get("matched_external_key"):
                platform_code = raw["matched_platform"]
                external_key = raw["matched_external_key"]
                matched_location = location_index.get((platform_code, external_key)) or location_index.get(
                    (platform_code, norm_slug(external_key))
                )

            runpark_location_row = None
            if raw.get("show_on_map") and raw.get("runpark_slug"):
                slug = raw["runpark_slug"]
                display_name = raw.get("display_name") or raw["runpark_name"]
                runpark_location_row = location_index.get(("runpark", slug))
                is_paused = bool(raw.get("is_paused", False))
                if runpark_location_row is None:
                    runpark_location_row = Location(
                        platform_id=runpark_platform.id,
                        external_key=slug,
                        name=display_name,
                        country=raw.get("country"),
                        city=raw.get("city"),
                        region=raw.get("region"),
                        latitude=raw.get("latitude"),
                        longitude=raw.get("longitude"),
                        is_official_map=True,
                        is_paused=is_paused,
                        source_url=raw.get("public_url"),
                    )
                    db.add(runpark_location_row)
                    db.flush()
                    location_index[("runpark", slug)] = runpark_location_row
                    stats["runpark_locations"] += 1
                else:
                    runpark_location_row.name = display_name
                    runpark_location_row.country = raw.get("country") or runpark_location_row.country
                    runpark_location_row.city = raw.get("city") or runpark_location_row.city
                    runpark_location_row.region = raw.get("region") or runpark_location_row.region
                    runpark_location_row.latitude = raw.get("latitude") or runpark_location_row.latitude
                    runpark_location_row.longitude = raw.get("longitude") or runpark_location_row.longitude
                    runpark_location_row.is_official_map = True
                    runpark_location_row.is_paused = is_paused
                    runpark_location_row.source_url = raw.get("public_url") or runpark_location_row.source_url

            mapping = RunparkLocationMapping(
                runpark_location_id=raw["runpark_location_id"],
                runpark_slug=raw.get("runpark_slug"),
                runpark_name=raw["runpark_name"],
                display_name=raw.get("display_name"),
                decision=raw["decision"],
                show_on_map=bool(raw.get("show_on_map")),
                is_transitional=bool(raw.get("is_transitional")),
                is_paused=bool(raw.get("is_paused", False)),
                latitude=raw.get("latitude"),
                longitude=raw.get("longitude"),
                city=raw.get("city"),
                region=raw.get("region"),
                country=raw.get("country"),
                matched_platform=raw.get("matched_platform"),
                matched_external_key=raw.get("matched_external_key"),
                matched_location_id=matched_location.id if matched_location else None,
                legacy_parkrun_slug=raw.get("legacy_parkrun_slug"),
                runpark_location_row_id=runpark_location_row.id if runpark_location_row else None,
                public_url=raw.get("public_url"),
                duplicate_match_text=raw.get("duplicate_match_text"),
                decision_raw=raw.get("decision_raw"),
                source_batch=raw["source_batch"],
                notes=raw.get("notes"),
            )
            db.add(mapping)
            stats["mappings"] += 1
            if matched_location:
                stats["matched_locations"] += 1

            _sync_catalog_links(
                db,
                raw,
                platforms,
                location_index,
                catalog_by_slug,
                runpark_location_row,
                matched_location,
                stats,
            )

        db.commit()
        return stats
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def _ensure_catalog_link(
    db,
    *,
    catalog: LocationCatalog,
    platform: Platform,
    external_key: str,
    location: Location | None,
    stats: dict[str, int],
) -> None:
    from app.models import LocationCatalogLink

    existing = (
        db.query(LocationCatalogLink)
        .filter(
            LocationCatalogLink.platform_id == platform.id,
            LocationCatalogLink.external_key == external_key,
        )
        .one_or_none()
    )
    if existing is not None:
        if location is not None and existing.location_id is None:
            existing.location_id = location.id
        if existing.catalog_id != catalog.id and existing.location_id is None and location is not None:
            existing.catalog_id = catalog.id
        return

    db.add(
        LocationCatalogLink(
            catalog_id=catalog.id,
            platform_id=platform.id,
            external_key=external_key,
            location_id=location.id if location else None,
        )
    )
    stats["catalog_links"] += 1


def _sync_catalog_links(
    db,
    raw: dict,
    platforms: dict,
    location_index: dict,
    catalog_by_slug: dict,
    runpark_location_row,
    matched_location,
    stats: dict[str, int],
) -> None:
    from app.models import LocationCatalog
    from app.runpark.mappings import RunparkMappingDecision

    decision = raw["decision"]
    legacy_slug = raw.get("legacy_parkrun_slug")
    runpark_slug = raw.get("runpark_slug")
    display_name = raw.get("display_name") or raw["runpark_name"]

    catalog = catalog_by_slug.get(legacy_slug) if legacy_slug else None

    if decision in (RunparkMappingDecision.KEEP_RUNPARK, RunparkMappingDecision.MATCH_PARKRUN):
        if catalog is None:
            catalog = LocationCatalog(
                canonical_name=display_name,
                legacy_parkrun_slug=legacy_slug,
                active_platform="runpark",
                notes=raw.get("notes"),
            )
            db.add(catalog)
            db.flush()
            if legacy_slug:
                catalog_by_slug[legacy_slug] = catalog
        else:
            catalog.active_platform = "runpark"
            if display_name:
                catalog.canonical_name = display_name

        if runpark_slug and runpark_location_row is not None:
            _ensure_catalog_link(
                db,
                catalog=catalog,
                platform=platforms["runpark"],
                external_key=runpark_slug,
                location=runpark_location_row,
                stats=stats,
            )

        if decision == RunparkMappingDecision.MATCH_PARKRUN and legacy_slug:
            parkrun_location = location_index.get(("parkrun", legacy_slug))
            _ensure_catalog_link(
                db,
                catalog=catalog,
                platform=platforms["parkrun"],
                external_key=legacy_slug,
                location=parkrun_location,
                stats=stats,
            )
        return

    if decision == RunparkMappingDecision.TRANSITIONAL and legacy_slug:
        if catalog is None:
            catalog = LocationCatalog(
                canonical_name=display_name,
                legacy_parkrun_slug=legacy_slug,
                active_platform="five_verst",
                notes=raw.get("notes") or raw.get("decision_raw"),
            )
            db.add(catalog)
            db.flush()
            catalog_by_slug[legacy_slug] = catalog
        else:
            note = raw.get("notes") or ""
            if note and (catalog.notes or "").find(note) < 0:
                catalog.notes = "; ".join(filter(None, [catalog.notes, note]))

        if runpark_slug:
            _ensure_catalog_link(
                db,
                catalog=catalog,
                platform=platforms["runpark"],
                external_key=runpark_slug,
                location=None,
                stats=stats,
            )
        if matched_location is not None:
            platform_code = raw.get("matched_platform")
            external_key = raw.get("matched_external_key")
            if platform_code and external_key:
                _ensure_catalog_link(
                    db,
                    catalog=catalog,
                    platform=platforms[platform_code],
                    external_key=external_key,
                    location=matched_location,
                    stats=stats,
                )
        parkrun_location = location_index.get(("parkrun", legacy_slug))
        if parkrun_location is not None:
            _ensure_catalog_link(
                db,
                catalog=catalog,
                platform=platforms["parkrun"],
                external_key=legacy_slug,
                location=parkrun_location,
                stats=stats,
            )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--import-db", action="store_true")
    parser.add_argument("--from-json", action="store_true")
    parser.add_argument("--geocode", action="store_true", help="Fill country/region/city via Nominatim")
    args = parser.parse_args()

    if args.from_json:
        entries = load_json(args.json)
    else:
        if backend_path := str(BACKEND_ROOT):
            if backend_path not in sys.path:
                sys.path.insert(0, backend_path)
        entries = load_xlsx(args.xlsx)
        export_json(entries, args.json)

    if args.geocode:
        entries = geocode_entries(entries)

    print(f"Entries: {len(entries)}")
    print(f"  keep_runpark: {sum(1 for e in entries if e['decision'] == 'keep_runpark')}")
    print(f"  match_parkrun: {sum(1 for e in entries if e['decision'] == 'match_parkrun')}")
    print(f"  remove: {sum(1 for e in entries if e['decision'] == 'remove_from_runpark')}")
    print(f"  transitional: {sum(1 for e in entries if e['decision'] == 'transitional')}")
    print(f"  show_on_map: {sum(1 for e in entries if e.get('show_on_map'))}")
    print(f"JSON: {args.json}")

    if args.geocode and not args.from_json:
        export_json(entries, args.json)

    if args.import_db:
        stats = import_to_db(entries)
        print(f"DB import: {stats}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Build parkrun → 5verst/s95 mapping XLSX for manual review."""

from __future__ import annotations

import argparse
import json
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = ROOT / "data" / "location_mapping_draft.xlsx"

PARKRUN_SLUG_ALIASES: dict[str, str] = {
    "gorky-park": "gorkypark",
    "park-850-letiya-moskvy": "park850letiyamoskvy",
    "tula-central": "tulacentral",
    "tver-rechnoy-vokzal": "tverrechnoyvokzal",
    "moskovsky-park-pobedy": "moskovskyparkpobedy",
    "lobnya-gorodskoy-park": "lobnyagorodskoypark",
    "serpukhov-gorodskoy-bor": "serpukhovgorodskoybor",
    "gatchina-prioratsky": "gatchinaprioratsky",
    "balashikha-zarechnaya": "balashikhazarechnaya",
    "babushkinsky-na-yauze": "babushkinskynayauze",
    "pokrovskoe-streshnevo": "pokrovskoestreshnevo",
    "severnoe-tushino": "severnoetushino",
    "olimpiyskaya-derevnya": "olimpiyskayaderevnya",
    "lesopark-severny": "lesoparkseverny",
}

PARKRUN_TO_S95_SLUG: dict[str, str] = {
    "izmailovo": "izmailovo",
    "tsaritsyno": "tsaritsyno",
    "olimpiyskaya-derevnya": "olimpiyskaya_derevnya",
    "tver-rechnoy-vokzal": "parkzhrun",
    "gatchina-prioratsky": "gatchina",
}

PARKRUN_DISPLAY_RU: dict[str, str] = {
    "lesopark-severny": "Лесопарк Северный",
    "pokrovskoe-streshnevo": "Покровское-Стрешнево",
    "tver-rechnoy-vokzal": "Тверь",
}


@dataclass
class LocRow:
    external_key: str
    name: str


def norm_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def norm_token(value: str) -> str:
    value = value.lower().replace("ё", "е")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def extract_5verst_slug(url: str | None) -> str:
    if not url:
        return ""
    match = re.search(r"5verst\.ru/([^/]+)/?", url)
    return match.group(1) if match else ""


def extract_s95_slug(url: str | None) -> str:
    if not url:
        return ""
    match = re.search(r"s95\.ru/events/([^/\"?#]+)", url)
    return match.group(1) if match else ""


def grafana_rows(payload: dict) -> list[tuple]:
    frames = payload.get("results", {}).get("A", {}).get("frames", [])
    if not frames:
        return []
    values = frames[0].get("data", {}).get("values", [])
    if not values:
        return []
    return list(zip(*values))


def fetch_run5k(sql: str) -> dict:
    subprocess.run(
        [
            "curl", "-sS", "-c", "/tmp/grafana_cookies.txt",
            "-X", "POST", "https://run5k.run/login",
            "-H", "Content-Type: application/json",
            "-d", '{"user":"user","password":"user"}',
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    proc = subprocess.run(
        [
            "curl", "-sS", "-b", "/tmp/grafana_cookies.txt",
            "-X", "POST", "https://run5k.run/api/ds/query",
            "-H", "Content-Type: application/json",
            "-d",
            json.dumps(
                {
                    "queries": [{
                        "refId": "A",
                        "datasource": {"type": "grafana-postgresql-datasource", "uid": "de407hodrx0jkd"},
                        "rawSql": sql,
                        "format": "table",
                    }],
                    "from": "now-1h",
                    "to": "now",
                }
            ),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(proc.stdout)


def load_parkrun_from_db() -> tuple[dict[str, LocRow], dict[str, LocRow]]:
    sql = """
    SELECT l.external_key, l.name
    FROM locations l
    JOIN platforms p ON l.platform_id = p.id
    WHERE p.code = 'parkrun' AND l.external_key NOT IN ('parkrun-summary')
    ORDER BY l.name
    """
    proc = subprocess.run(
        [
            "docker", "compose", "exec", "-T", "postgres",
            "psql", "-U", "saturday_runs", "-d", "saturday_runs_lk",
            "-t", "-A", "-F|", "-c", sql,
        ],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    by_slug: dict[str, LocRow] = {}
    by_name: dict[str, LocRow] = {}
    for line in proc.stdout.splitlines():
        if not line.strip():
            continue
        slug, name = line.split("|", 1)
        row = LocRow(external_key=slug, name=name)
        by_slug[norm_slug(slug)] = row
        by_name[norm_token(name)] = row
    return by_slug, by_name


def load_run5k_catalog() -> tuple[dict[str, str], dict[str, dict]]:
    """five_verst slug (normalized) -> canonical Russian name; s95 slug -> info."""
    fv_by_slug: dict[str, str] = {}
    for name_point, link_point in grafana_rows(
        fetch_run5k("SELECT name_point, link_point FROM general_link_all_location ORDER BY name_point")
    ):
        slug = extract_5verst_slug(link_point)
        if slug:
            fv_by_slug[norm_slug(slug)] = name_point

    s95_by_slug: dict[str, dict] = {}
    for short_name, full_name, link_point in grafana_rows(
        fetch_run5k("SELECT name_point, full_name_point, link_point FROM s95_location ORDER BY name_point")
    ):
        slug = extract_s95_slug(link_point)
        if slug:
            s95_by_slug[norm_slug(slug)] = {
                "name": short_name,
                "full_name": full_name or short_name,
                "url": link_point or "",
            }
    return fv_by_slug, s95_by_slug


def parkrun_slug_for_name(parkrun_name: str, db_by_name: dict[str, LocRow]) -> str:
    row = db_by_name.get(norm_token(parkrun_name))
    if row:
        return row.external_key
    # fallback: derive slug-like key from english name
    return norm_slug(parkrun_name.replace(" ", "-"))


def resolve_mapping(
    parkrun_name: str,
    db_by_slug: dict[str, LocRow],
    db_by_name: dict[str, LocRow],
    fv_by_slug: dict[str, str],
    s95_by_slug: dict[str, dict],
) -> dict:
    slug = parkrun_slug_for_name(parkrun_name, db_by_name)
    fv_key = PARKRUN_SLUG_ALIASES.get(slug, norm_slug(slug))

    if fv_key in fv_by_slug:
        return {
            "parkrun_location": parkrun_name,
            "current_location": fv_by_slug[fv_key],
            "current_system": "5 вёрст",
            "comment": f"Сопоставлено по slug parkrun → 5verst ({slug} → {fv_key}). Источник: run5k general_link.",
            "confirmed": "",
            "parkrun_slug": slug,
            "actual_slug": fv_key,
            "match_confidence": "высокая",
        }

    s95_key = norm_slug(PARKRUN_TO_S95_SLUG.get(slug, slug))
    s95 = s95_by_slug.get(s95_key)
    if s95:
        return {
            "parkrun_location": parkrun_name,
            "current_location": s95["name"],
            "current_system": "С95",
            "comment": (
                f"В каталоге 5verst (run5k) нет — вероятно перешло на С95 ({s95_key}). "
                f"Полное название: {s95['full_name']}."
            ),
            "confirmed": "",
            "parkrun_slug": slug,
            "actual_slug": s95_key,
            "match_confidence": "средняя",
        }

    if slug in PARKRUN_DISPLAY_RU:
        return {
            "parkrun_location": parkrun_name,
            "current_location": PARKRUN_DISPLAY_RU[slug],
            "current_system": "нет (только parkrun)",
            "comment": "Не найдено в run5k (5verst/s95). Вероятно закрытая или переименованная локация — уточнить вручную.",
            "confirmed": "",
            "parkrun_slug": slug,
            "actual_slug": "",
            "match_confidence": "низкая",
        }

    return {
        "parkrun_location": parkrun_name,
        "current_location": "",
        "current_system": "",
        "comment": f"Автосопоставление не найдено (parkrun slug: {slug}). Заполните вручную.",
        "confirmed": "",
        "parkrun_slug": slug,
        "actual_slug": "",
        "match_confidence": "нет",
    }


def build_rows() -> list[dict]:
    db_by_slug, db_by_name = load_parkrun_from_db()
    fv_by_slug, s95_by_slug = load_run5k_catalog()

    parkrun_names = [
        row[0]
        for row in grafana_rows(
            fetch_run5k("SELECT name_point FROM parkrun_russia_location ORDER BY name_point")
        )
        if row and row[0]
    ]

    rows = [
        resolve_mapping(name, db_by_slug, db_by_name, fv_by_slug, s95_by_slug)
        for name in parkrun_names
    ]

    # Ensure every parkrun slug from our DB appears even if missing in run5k list
    known_names = {norm_token(r["parkrun_location"]) for r in rows}
    for slug_norm, loc in db_by_slug.items():
        if norm_token(loc.name) in known_names:
            continue
        rows.append(
            resolve_mapping(loc.name, db_by_slug, db_by_name, fv_by_slug, s95_by_slug)
        )

    rows.sort(key=lambda r: r["parkrun_location"].lower())
    return rows


def write_xlsx(path: Path, rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Main sheet: 5 required columns ---
    ws = wb.active
    ws.title = "соответствия"
    main_headers = [
        ("Локация parkrun", "Историческое название локации в parkrun Russia (англ. имя из run5k)."),
        ("Актуальная локация", "Как называется сейчас в С95 или 5 вёрст. Пусто — заполнить вручную."),
        ("Система", "Где актуальна: «5 вёрст», «С95» или «нет (только parkrun)»."),
        ("Комментарий", "Пояснение автоматического сопоставления и что проверить."),
        ("Подтверждение", "Ваше поле: «да», «нет» или исправленное значение. Верните файл после проверки."),
    ]
    extra_headers = [
        ("parkrun_slug", "Техн. slug parkrun в нашей БД (для импорта)."),
        ("actual_slug", "Техн. slug актуальной локации в 5verst.ru или s95.ru."),
        ("уверенность", "Оценка автосопоставления: высокая / средняя / низкая / нет."),
    ]

    all_headers = [h[0] for h in main_headers + extra_headers]
    ws.append(all_headers)

    for row in rows:
        ws.append([
            row["parkrun_location"],
            row["current_location"],
            row["current_system"],
            row["comment"],
            row["confirmed"],
            row["parkrun_slug"],
            row["actual_slug"],
            row["match_confidence"],
        ])

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    main_fill = PatternFill("solid", fgColor="E8F4FC")
    for col, (title, _desc) in enumerate(main_headers + extra_headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill if col <= 5 else main_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        width = 36 if col in {1, 2, 4} else 16
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(rows) + 1}"

    for row_idx in range(2, len(rows) + 2):
        ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    # --- Column documentation ---
    doc = wb.create_sheet("описание столбцов")
    doc.append(["Столбец", "Лист", "Что означает"])
    doc.column_dimensions["A"].width = 22
    doc.column_dimensions["B"].width = 14
    doc.column_dimensions["C"].width = 72
    for title, desc in main_headers + extra_headers:
        sheet_name = "соответствия"
        doc.append([title, sheet_name, desc])
    doc.append([])
    doc.append(["Источники данных", "", ""])
    doc.append(["parkrun_russia_location", "run5k.run", "Полный список локаций parkrun Russia (105)"])
    doc.append(["general_link_all_location", "run5k.run", "Канонические русские имена → URL 5verst.ru"])
    doc.append(["s95_location", "run5k.run", "Локации С95"])
    doc.append(["locations (parkrun)", "наша БД", "Slug parkrun из импортированных профилей"])
    doc.append([])
    doc.append(["Как проверять", "", ""])
    doc.append([
        "1",
        "",
        "Проверьте столбцы «Актуальная локация» и «Система». Исправьте при необходимости.",
    ])
    doc.append([
        "2",
        "",
        "В «Подтверждение» напишите «да» или укажите правильное значение.",
    ])
    doc.append([
        "3",
        "",
        "Строки с «нет (только parkrun)» или пустой «Системой» — закрытые/спорные, нужен ручной ввод.",
    ])
    doc.append([])
    doc.append(["Статистика", "", f"Всего строк parkrun: {len(rows)}"])
    doc.append([
        "",
        "",
        f"5 вёрст: {sum(1 for r in rows if r['current_system'] == '5 вёрст')}",
    ])
    doc.append([
        "",
        "",
        f"С95: {sum(1 for r in rows if r['current_system'] == 'С95')}",
    ])
    doc.append([
        "",
        "",
        f"Только parkrun / без пары: {sum(1 for r in rows if r['current_system'] in ('', 'нет (только parkrun)'))}",
    ])

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    rows = build_rows()
    write_xlsx(args.output, rows)

    print(f"Wrote {args.output}")
    print(f"  parkrun rows: {len(rows)}")
    print(f"  → 5 вёрст: {sum(1 for r in rows if r['current_system'] == '5 вёрст')}")
    print(f"  → С95: {sum(1 for r in rows if r['current_system'] == 'С95')}")
    print(f"  → без пары: {sum(1 for r in rows if not r['current_system'] or r['current_system'] == 'нет (только parkrun)')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

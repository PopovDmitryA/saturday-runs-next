from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

runpark_locations = [
    {"name": "RUNtonda Mikhalkovo (Москва)", "city": "Москва", "event_count": 371, "first_event_date": "2022-03-12", "last_event_date": "2026-06-13", "public_url": "https://runpark.ru/Places/RUNtonda-Mikhalkovo"},
    {"name": "Великий Новгород", "city": "Великий Новгород", "event_count": 220, "first_event_date": "2022-03-19", "last_event_date": "2026-06-20", "public_url": "https://runpark.ru/Places/velikiynovgorodkremlevsky"},
    {"name": "Парк Ангарские Пруды", "city": "Москва", "event_count": 230, "first_event_date": "2022-03-12", "last_event_date": "2026-06-20", "public_url": "https://runpark.ru/Places/angarskie-prudy"},
    {"name": "Лесопарк «Ублинские горы» (ходьба)", "city": "Старый Оскол", "event_count": 12, "first_event_date": "2025-10-11", "last_event_date": "2026-02-21", "public_url": "https://runpark.ru/Places/oskol_ublinskie_gory"},
    {"name": "Парк Победы 5 км", "city": "Белгород", "event_count": 180, "first_event_date": "2022-05-21", "last_event_date": "2026-01-10", "public_url": "https://runpark.ru/Places/belgorodparkpobedy"},
    {"name": "Kstovo5RUN", "city": "Кстово (Нижний Новгород)", "event_count": 37, "first_event_date": "2025-04-27", "last_event_date": "2026-01-10", "public_url": "https://runpark.ru/Places/Kstovo5run"},
    {"name": "Парк Победы, суббота, 9-00, 5 км", "city": "Липецк", "event_count": 126, "first_event_date": "2023-07-08", "last_event_date": "2025-12-20", "public_url": "https://runpark.ru/Places/Lipetsk_parkpobeda"},
    {"name": "5кмПритяжения", "city": "Магнитогорск", "event_count": 26, "first_event_date": "2024-04-13", "last_event_date": "2025-11-19", "public_url": "https://runpark.ru/Places/magnitprityagenie"},
    {"name": "паркбег40", "city": "Калуга", "event_count": 133, "first_event_date": "2022-04-02", "last_event_date": "2025-05-24", "public_url": "https://runpark.ru/Places/parkbeg40"},
    {"name": "СормовоРан", "city": "Нижний Новгород", "event_count": 23, "first_event_date": "2024-09-07", "last_event_date": "2025-05-03", "public_url": "https://runpark.ru/Places/sormovo"},
    {"name": "GreenRun Киров", "city": "Киров", "event_count": 26, "first_event_date": "2022-03-19", "last_event_date": "2024-10-26", "public_url": "https://runpark.ru/Places/greenrunkirov"},
    {"name": "Южный парк", "city": "Магнитогорск", "event_count": 1, "first_event_date": "2024-10-26", "last_event_date": "2024-10-26", "public_url": "https://runpark.ru/Places/southparknight"},
    {"name": "Беговая суббота / Курск", "city": "Курск", "event_count": 14, "first_event_date": "2022-03-26", "last_event_date": "2024-05-04", "public_url": "https://runpark.ru/Places/boevadacha"},
    {"name": "ШвейцRUN", "city": "Нижний Новгород", "event_count": 24, "first_event_date": "2023-02-05", "last_event_date": "2023-07-15", "public_url": "https://runpark.ru/Places/swiss5run"},
    {"name": "Парк Горького", "city": "Москва", "event_count": 1, "first_event_date": "2023-04-15", "last_event_date": "2023-04-15", "public_url": "https://runpark.ru/Places/gorkypark"},
    {"name": "Дружба", "city": "Москва", "event_count": 24, "first_event_date": "2022-03-12", "last_event_date": "2023-03-04", "public_url": "https://runpark.ru/Places/druzhba"},
    {"name": "BGO.RUN / Ногинский городской парк", "city": "Ногинск", "event_count": 29, "first_event_date": "2022-03-19", "last_event_date": "2022-10-29", "public_url": "https://runpark.ru/Places/runbogorodsk"},
    {"name": "Балашиха Заречная", "city": "Балашиха", "event_count": 19, "first_event_date": "2022-03-12", "last_event_date": "2022-07-23", "public_url": "https://runpark.ru/Places/balashikhazarechnaya"},
    {"name": "Крылатское", "city": "Москва", "event_count": 14, "first_event_date": "2022-03-19", "last_event_date": "2022-06-18", "public_url": "https://runpark.ru/Places/krylatskoe"},
    {"name": "Забег Приятелей / Buddy Race", "city": "Железнодорожный", "event_count": 12, "first_event_date": "2022-03-26", "last_event_date": "2022-06-11", "public_url": "https://runpark.ru/Places/pestovskiypark"},
    {"name": "Тимирязевский парк", "city": "Москва", "event_count": 13, "first_event_date": "2022-03-12", "last_event_date": "2022-06-04", "public_url": "https://runpark.ru/Places/timiryazevsky"},
    {"name": "Химки", "city": "Химки", "event_count": 12, "first_event_date": "2022-03-19", "last_event_date": "2022-06-04", "public_url": "https://runpark.ru/Places/khimki"},
    {"name": "Уфа Лесопарк Лесоводов", "city": "Уфа", "event_count": 11, "first_event_date": "2022-03-26", "last_event_date": "2022-06-04", "public_url": "https://runpark.ru/Places/parklesovodov"},
    {"name": "Readovsky Park", "city": "Смоленск", "event_count": 12, "first_event_date": "2022-03-19", "last_event_date": "2022-06-04", "public_url": "https://runpark.ru/Places/readovskypark"},
    {"name": "Butovo Run", "city": "Москва", "event_count": 11, "first_event_date": "2022-03-19", "last_event_date": "2022-06-04", "public_url": "https://runpark.ru/Places/butovo"},
    {"name": "Бег в парке Железноводск", "city": "Железноводск", "event_count": 2, "first_event_date": "2022-03-26", "last_event_date": "2022-06-04", "public_url": "https://runpark.ru/Places/zeleznovodsk"},
    {"name": "Нижний парк, Пушкин", "city": "Санкт-Петербург", "event_count": 5, "first_event_date": "2022-03-26", "last_event_date": "2022-05-28", "public_url": "https://runpark.ru/Places/pushkin"},
    {"name": "Муромец", "city": "Муром", "event_count": 10, "first_event_date": "2022-03-19", "last_event_date": "2022-05-28", "public_url": "https://runpark.ru/Places/muromets"},
    {"name": "Сосновка", "city": "Санкт-Петербург", "event_count": 9, "first_event_date": "2022-03-19", "last_event_date": "2022-05-21", "public_url": "https://runpark.ru/Places/sosnovka"},
    {"name": "Бег в парке", "city": "Ставрополь", "event_count": 9, "first_event_date": "2022-03-26", "last_event_date": "2022-05-21", "public_url": "https://runpark.ru/Places/stavropol"},
    {"name": "Парк Строителей Орск", "city": "Орск", "event_count": 9, "first_event_date": "2022-03-26", "last_event_date": "2022-05-21", "public_url": "https://runpark.ru/Places/orskparkstroiteley"},
    {"name": "Воркута Городской парк", "city": "Воркута", "event_count": 10, "first_event_date": "2022-03-19", "last_event_date": "2022-05-21", "public_url": "https://runpark.ru/Places/vorkutagorodskoypark"},
    {"name": "Парковые забеги ЖукRun", "city": "Жуковский", "event_count": 9, "first_event_date": "2022-03-26", "last_event_date": "2022-05-21", "public_url": "https://runpark.ru/Places/zhukovsky"},
    {"name": "ZабегVборуSерпухов", "city": "Серпухов", "event_count": 5, "first_event_date": "2022-03-19", "last_event_date": "2022-05-14", "public_url": "https://runpark.ru/Places/serpukhov"},
    {"name": "Парковый бег, Раменское", "city": "Раменское", "event_count": 9, "first_event_date": "2022-03-19", "last_event_date": "2022-05-14", "public_url": "https://runpark.ru/Places/ramenskoegorodskoypark"},
    {"name": "Казань Парк Горького", "city": "Казань", "event_count": 2, "first_event_date": "2022-03-19", "last_event_date": "2022-05-14", "public_url": "https://runpark.ru/Places/kazancentral"},
    {"name": "Mytischi Be Good", "city": "Мытищи", "event_count": 5, "first_event_date": "2022-03-19", "last_event_date": "2022-04-30", "public_url": "https://runpark.ru/Places/mytishchicentralpark"},
    {"name": "Чертаново", "city": "Москва", "event_count": 4, "first_event_date": "2022-03-19", "last_event_date": "2022-04-16", "public_url": "https://runpark.ru/Places/chertanovopokrovskypark"},
    {"name": "MyRUNo — парк 850-летия (Марьино)", "city": "Москва", "event_count": 5, "first_event_date": "2022-03-12", "last_event_date": "2022-04-09", "public_url": "https://runpark.ru/Places/myRUNo"},
    {"name": "runкурган", "city": "Курган", "event_count": 1, "first_event_date": "2022-03-26", "last_event_date": "2022-03-26", "public_url": "https://runpark.ru/Places/runkurgan"},
]

parkrun_no_successor = [
    {"name": "Angarskie Prudy", "slug": "angarskie-prudy", "event_count": 23, "last_event": "2022-02-26"},
    {"name": "Kuzminki", "slug": "kuzminki", "event_count": 108, "last_event": "2022-02-26"},
    {"name": "Natashinsky", "slug": "natashinsky", "event_count": 83, "last_event": "2022-02-26"},
    {"name": "Butovo", "slug": "butovo", "event_count": 14, "last_event": "2022-02-19"},
    {"name": "Ramenskoe Gorodskoy Park", "slug": "ramenskoe-gorodskoy-park", "event_count": 6, "last_event": "2022-02-19"},
    {"name": "Sosnovka", "slug": "sosnovka", "event_count": 4, "last_event": "2022-02-19"},
    {"name": "Velikiy Novgorod Kremlevsky", "slug": "velikiy-novgorod-kremlevsky", "event_count": 6, "last_event": "2022-02-12"},
    {"name": "park 30-letiya Oktyabrya", "slug": "park-30-letiya-oktyabrya", "event_count": 3, "last_event": "2022-02-12"},
    {"name": "Mytishchi Central Park", "slug": "mytishchi-central-park", "event_count": 4, "last_event": "2022-01-15"},
    {"name": "Pushkin", "slug": "pushkin", "event_count": 2, "last_event": "2022-01-08"},
    {"name": "Internatsionalistov", "slug": "internatsionalistov", "event_count": 2, "last_event": "2022-01-07"},
    {"name": "Petergof Aleksandriysky", "slug": "petergof-aleksandriysky", "event_count": 30, "last_event": "2021-12-18"},
    {"name": "Nizhny Novgorod Meshchersky", "slug": "nizhny-novgorod-meshchersky", "event_count": 3, "last_event": "2021-12-11"},
    {"name": "Chertanovo Pokrovsky Park", "slug": "chertanovo-pokrovsky-park", "event_count": 7, "last_event": "2021-12-04"},
    {"name": "Ulitsa Golubaya", "slug": "ulitsa-golubaya", "event_count": 3, "last_event": "2021-12-04"},
    {"name": "Kimry", "slug": "kimry", "event_count": 5, "last_event": "2021-11-06"},
    {"name": "Mega Park Kudrovo", "slug": "mega-park-kudrovo", "event_count": 4, "last_event": "2021-10-30"},
    {"name": "Korolev", "slug": "korolev", "event_count": 7, "last_event": "2021-10-23"},
    {"name": "Rostov on Don", "slug": "rostov-on-don", "event_count": 2, "last_event": "2021-10-23"},
    {"name": "Filatov Lug", "slug": "filatov-lug", "event_count": 9, "last_event": "2021-10-09"},
    {"name": "Pokrovskoe-Streshnevo", "slug": "pokrovskoe-streshnevo", "event_count": 16, "last_event": "2021-10-02"},
    {"name": "Park Talalikhina", "slug": "park-talalikhina", "event_count": 3, "last_event": "2021-09-11"},
    {"name": "Solnechny ostrov", "slug": "solnechny-ostrov", "event_count": 2, "last_event": "2021-09-04"},
    {"name": "Starye Sady", "slug": "starye-sady", "event_count": 1, "last_event": "2021-08-28"},
    {"name": "Naberezhnaya Reki Velikoy", "slug": "naberezhnaya-reki-velikoy", "event_count": 2, "last_event": "2021-08-21"},
    {"name": "Tambov", "slug": "tambov", "event_count": 5, "last_event": "2021-08-07"},
    {"name": "Samara Park Gagarina", "slug": "samara-park-gagarina", "event_count": 1, "last_event": "2021-07-31"},
    {"name": "Memorialny Park", "slug": "memorialny-park", "event_count": 1, "last_event": "2021-07-24"},
    {"name": "Kolpino", "slug": "kolpino", "event_count": 2, "last_event": "2021-07-10"},
    {"name": "Nizhny Prud", "slug": "nizhny-prud", "event_count": 1, "last_event": "2021-07-03"},
    {"name": "Shuvalovsky Park", "slug": "shuvalovsky-park", "event_count": 2, "last_event": "2021-06-05"},
    {"name": "Elagin Ostrov", "slug": "elagin-ostrov", "event_count": 5, "last_event": "2021-05-29"},
    {"name": "Gorodskoy Park Semya", "slug": "gorodskoy-park-semya", "event_count": 1, "last_event": "2021-05-22"},
    {"name": "Yarmarochnaya Ploshchad", "slug": "yarmarochnaya-ploshchad", "event_count": 1, "last_event": "2021-05-22"},
    {"name": "Maykop Gorodskoy Park", "slug": "maykop-gorodskoy-park", "event_count": 1, "last_event": "2021-05-15"},
    {"name": "Skver Dzerzhinskogo", "slug": "skver-dzerzhinskogo", "event_count": 1, "last_event": "2021-05-15"},
    {"name": "Petrovskaya naberezhnaya", "slug": "petrovskaya-naberezhnaya", "event_count": 1, "last_event": "2021-05-08"},
    {"name": "Kazan Central", "slug": "kazan-central", "event_count": 2, "last_event": "2021-05-01"},
    {"name": "Mitino", "slug": "mitino", "event_count": 7, "last_event": "2021-04-17"},
    {"name": "Noginsk Gorodskoy Park", "slug": "noginsk-gorodskoy-park", "event_count": 3, "last_event": "2021-04-10"},
    {"name": "Stavropol", "slug": "stavropol", "event_count": 1, "last_event": "2021-03-27"},
    {"name": "Yoshkar-Ola Alleya Zdorovya", "slug": "yoshkar-ola-alleya-zdorovya", "event_count": 1, "last_event": "2021-03-13"},
    {"name": "Boeva Dacha", "slug": "boeva-dacha", "event_count": 1, "last_event": "2021-01-02"},
    {"name": "Krasnoyarsk Naberezhnaya", "slug": "krasnoyarsk-naberezhnaya", "event_count": 1, "last_event": "2020-12-19"},
    {"name": "Zhukovsky", "slug": "zhukovsky", "event_count": 5, "last_event": "2020-02-08"},
    {"name": "Kurgan Central Park", "slug": "kurgan-central-park", "event_count": 1, "last_event": "2020-01-11"},
    {"name": "Readovsky Park", "slug": "readovsky-park", "event_count": 2, "last_event": "2020-01-04"},
]

wb = Workbook()

# === Лист 1: RunPark локации (40 неподключённых) ===
ws1 = wb.active
ws1.title = "RunPark — неподключённые"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
yellow_fill = PatternFill("solid", start_color="FFFACD")

headers = ["#", "Название (RunPark)", "Город", "Пробежек", "Первая пробежка", "Последняя пробежка", "Ссылка RunPark", "Комментарий (тип перехода)"]
col_widths = [4, 42, 22, 10, 16, 18, 48, 55]

for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws1.cell(row=1, column=col_i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    ws1.column_dimensions[get_column_letter(col_i)].width = w

ws1.row_dimensions[1].height = 32

for row_i, loc in enumerate(runpark_locations, start=2):
    row_data = [
        row_i - 1,
        loc["name"],
        loc["city"],
        loc["event_count"],
        loc["first_event_date"],
        loc["last_event_date"],
        loc["public_url"],
        "",
    ]
    for col_i, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_i, column=col_i, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        if col_i == 8:
            cell.fill = yellow_fill
        if col_i in (5, 6):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_i == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_i == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Подсветить активные (последняя >= 2024)
    if loc["last_event_date"] >= "2024":
        for col_i in range(1, 8):
            ws1.cell(row=row_i, column=col_i).fill = PatternFill("solid", start_color="E8F5E9")

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:H{len(runpark_locations)+1}"

# Легенда внизу
legend_row = len(runpark_locations) + 3
ws1.cell(row=legend_row, column=1, value="Легенда:").font = Font(name="Arial", bold=True, size=10)
ws1.cell(row=legend_row+1, column=1, value="Зелёный фон").font = Font(name="Arial", size=10)
ws1.cell(row=legend_row+1, column=1).fill = PatternFill("solid", start_color="E8F5E9")
ws1.cell(row=legend_row+1, column=2, value="— Активная локация (пробежки после 2024 года)").font = Font(name="Arial", size=10)
ws1.cell(row=legend_row+2, column=1, value="Жёлтый фон").font = Font(name="Arial", size=10)
ws1.cell(row=legend_row+2, column=1).fill = yellow_fill
ws1.cell(row=legend_row+2, column=2, value="— Столбец для заполнения вручную").font = Font(name="Arial", size=10)

legend_types = [
    "Варианты для столбца «Комментарий»:",
    "  только RunPark — локация существует только в RunPark",
    "  parkrun → RunPark — была в parkrun, перешла в RunPark",
    "  parkrun → RunPark → 5verst — переход через RunPark в 5verst",
    "  parkrun → RunPark → s95 — переход через RunPark в s95",
    "  дубль 5verst — те же пробежки учтены в 5verst (основная)",
    "  дубль s95 — те же пробежки учтены в s95 (основная)",
    "  устаревшая — закрытая локация, не нужно подключать",
]
for i, txt in enumerate(legend_types):
    ws1.cell(row=legend_row+4+i, column=1, value=txt).font = Font(name="Arial", size=9, italic=True, color="555555")
    ws1.merge_cells(start_row=legend_row+4+i, start_column=1, end_row=legend_row+4+i, end_column=5)

# === Лист 2: Parkrun без преемников ===
ws2 = wb.create_sheet("Parkrun — без преемников")

for col_i, (h, w) in enumerate(zip(
    ["#", "Название (parkrun slug)", "Пробежек", "Последняя пробежка", "Преемник в RunPark?", "Комментарий"],
    [4, 38, 10, 16, 38, 50]
), start=1):
    cell = ws2.cell(row=1, column=col_i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    ws2.column_dimensions[get_column_letter(col_i)].width = w

ws2.row_dimensions[1].height = 32

for row_i, loc in enumerate(parkrun_no_successor, start=2):
    row_data = [row_i - 1, loc["name"], loc["event_count"], loc["last_event"], "", ""]
    for col_i, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_i, column=col_i, value=val)
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        if col_i in (5, 6):
            cell.fill = yellow_fill
        if col_i in (1, 3, 4):
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:F{len(parkrun_no_successor)+1}"

OUTPUT = "/Users/dmitry/Projects/saturday_runs_stats/docs/runpark/runpark_locations_mapping.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
